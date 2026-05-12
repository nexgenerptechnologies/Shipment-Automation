import frappe
from frappe.model.document import Document
from frappe.utils import nowdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Supplier Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Supplier Import"
    
    headers = [
        "Supplier ID (Leave blank for Auto)", "Naming Series", "Supplier Name", 
        "Supplier Group", "GST Category", "GSTIN",
        "Address Line 1", "City", "State", "Pincode", "Country",
        "Contact Person Name", "Email Address", "Mobile Number"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Supplier_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkSupplierImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_supplier_import.bulk_supplier_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_processing(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Suppliers, Addresses & Contacts. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_supplier_import.bulk_supplier_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "id": ["Supplier ID (Leave blank for Auto)", "Supplier ID"],
        "series": ["Naming Series"],
        "name": ["Supplier Name"],
        "group": ["Supplier Group"],
        "gst_cat": ["GST Category"],
        "gstin": ["GSTIN"],
        "street": ["Address Line 1", "Street Address"],
        "city": ["City"],
        "state": ["State"],
        "pincode": ["Pincode"],
        "country": ["Country"],
        "contact_name": ["Contact Person Name"],
        "email": ["Email Address"],
        "mobile": ["Mobile Number"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Supplier Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            manual_id = str(row[col_map["id"]]).strip() if col_map.get("id") is not None and row[col_map["id"]] else ""
            supplier_name = str(row[col_map["name"]]).strip() if col_map.get("name") is not None else ""
            gstin = str(row[col_map["gstin"]]).strip() if col_map.get("gstin") is not None and row[col_map["gstin"]] else ""
            gst_cat = str(row[col_map["gst_cat"]]).strip() if col_map.get("gst_cat") is not None else ""

            if not supplier_name:
                row_errors.append("Supplier Name is mandatory.")

            # Check Duplicates
            if manual_id and frappe.db.exists("Supplier", manual_id):
                row_errors.append(f"Manual ID '{manual_id}' already exists.")
            
            if not manual_id and frappe.db.exists("Supplier", supplier_name):
                row_errors.append(f"Supplier with name '{supplier_name}' already exists.")

            if gstin and frappe.db.exists("Supplier", {"gstin": gstin}):
                 row_errors.append(f"GSTIN '{gstin}' is already registered to another supplier.")

            if gst_cat and gst_cat not in frappe.get_meta("Supplier").get_field("gst_category").options.split("\n"):
                row_errors.append(f"Invalid GST Category: {gst_cat}")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Supplier Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        created = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            manual_id = str(row[col_map["id"]]).strip() if col_map.get("id") is not None and row[col_map["id"]] else ""
            series = str(row[col_map["series"]]).strip() if col_map.get("series") is not None and row[col_map["series"]] else ""
            supplier_name = str(row[col_map["name"]]).strip()
            group = str(row[col_map["group"]]).strip() if col_map.get("group") is not None and row[col_map["group"]] else ""
            
            # Check if group is a 'Group' type (ERPNext requires a non-group child for transactions)
            if group and frappe.db.get_value("Supplier Group", group, "is_group"):
                created.append(f"❌ {supplier_name}: Cannot select a Group type Supplier Group ({group}). Please select a non-group Supplier Group.")
                continue
            gst_cat = str(row[col_map["gst_cat"]]).strip() if col_map.get("gst_cat") is not None else ""
            gstin = str(row[col_map["gstin"]]).strip() if col_map.get("gstin") is not None and row[col_map["gstin"]] else ""
            
            try:
                # Use a transaction block (implicitly via insert/db_insert)
                # But we catch errors at each step and throw to prevent partial creation
                
                # 1. Create Supplier
                s_doc = frappe.new_doc("Supplier")
                
                # Naming Logic: Manual ID > Naming Series > Default
                if manual_id:
                    s_doc.name = manual_id
                    s_doc.flags.ignore_mandatory = True # To allow manual name setting
                elif series:
                    s_doc.naming_series = series
                
                s_doc.supplier_name = supplier_name
                s_doc.supplier_group = group
                s_doc.gst_category = gst_cat
                s_doc.gstin = gstin
                s_doc.insert(ignore_permissions=True)
                
                # Verification Check: Ensure name matches if manual_id was provided
                if manual_id and s_doc.name != manual_id:
                     # This happens if 'Buying Settings' is set to only use Supplier Name
                     # We force it by updating the name via database
                     frappe.db.set_value("Supplier", s_doc.name, "name", manual_id, update_modified=False)
                     s_doc.name = manual_id
                
                # 2. Create Address (Mandatory check: Line 1, State, Country)
                street = str(row[col_map["street"]]).strip() if col_map.get("street") is not None and row[col_map["street"]] else ""
                state = str(row[col_map["state"]]).strip() if col_map.get("state") is not None and row[col_map["state"]] else ""
                country = str(row[col_map["country"]]).strip() if col_map.get("country") is not None and row[col_map["country"]] else "India"

                if street or state: # If any address info is provided, enforce mandatory fields
                    if not street or not state or not country:
                        frappe.throw(f"Address creation failed: Address Line 1, State and Country are mandatory.")
                    
                    addr = frappe.new_doc("Address")
                    addr.address_title = supplier_name
                    addr.address_type = "Billing"
                    addr.address_line1 = street
                    addr.city = str(row[col_map["city"]]).strip() if col_map.get("city") is not None else ""
                    addr.state = state
                    
                    # Validate Pincode (6 digits, not starting with 0 for India)
                    pincode = str(row[col_map["pincode"]]).strip() if col_map.get("pincode") is not None else ""
                    if pincode and (len(pincode) != 6 or pincode.startswith("0")):
                         pincode = "" # Clear invalid pincode to prevent error
                    
                    addr.pincode = pincode
                    addr.country = country
                    addr.append("links", {"link_doctype": "Supplier", "link_name": s_doc.name})
                    addr.insert(ignore_permissions=True)

                # 3. Create Contact
                c_person = str(row[col_map["contact_name"]]).strip() if col_map.get("contact_name") is not None and row[col_map["contact_name"]] else ""
                email = str(row[col_map["email"]]).strip() if col_map.get("email") is not None and row[col_map["email"]] else ""
                mobile = str(row[col_map["mobile"]]).strip() if col_map.get("mobile") is not None and row[col_map["mobile"]] else ""

                if c_person:
                    con = frappe.new_doc("Contact")
                    con.first_name = c_person
                    con.append("links", {"link_doctype": "Supplier", "link_name": s_doc.name})
                    con.flags.ignore_permissions = True
                    con.db_insert()
                    
                    if email:
                         e_row = con.append("email_ids", {"email_id": email, "is_primary": 1})
                         e_row.db_insert()
                    
                    if mobile:
                         p_row = con.append("phone_nos", {"phone_number": mobile, "is_primary": 1})
                         p_row.db_insert()
                    
                    con.run_method("on_update")
                
                # 4. Final Verification Message
                msg = f"Supplier {supplier_name} created successfully"
                details = []
                if street or state: details.append("Address")
                if c_person: details.append("Contact")
                if email: details.append(f"Email({email})")
                if mobile: details.append(f"Mobile({mobile})")
                
                if details:
                    msg += f" with " + ", ".join(details)
                
                created.append(f"✅ {msg}")
            except Exception as e:
                # If anything fails, we rollback the specific supplier creation
                frappe.db.rollback()
                created.append(f"❌ {supplier_name}: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("processing_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
