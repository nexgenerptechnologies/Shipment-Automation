import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate
import openpyxl


class ShipmentImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        """Kick off background validation of the uploaded Excel file."""
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.shipment_import.shipment_import.run_validation",
            queue="long",
            timeout=3600,
            docname=self.name,
        )
        return "Validation started in the background. Refresh the page in a few seconds."

    @frappe.whitelist()
    def start_processing(self):
        """Kick off background processing: creates a single combined Purchase Receipt (Draft)."""
        if self.status != "Validated":
            frappe.throw("Please validate the data first before processing.")
        if not self.pr_naming_series:
            frappe.throw("Please select a Purchase Receipt Naming Series before processing.")
        self.db_set("status", "Processing")
        self.db_set("receipts_log", "⏳ Processing in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.shipment_import.shipment_import.run_processing",
            queue="long",
            timeout=3600,
            docname=self.name,
        )
        return "Processing started in the background. Refresh the page in a few seconds."

    @frappe.whitelist()
    def create_purchase_invoice_and_boe(self):
        """
        1. Creates a Purchase Invoice (Draft) from the submitted Purchase Receipt.
        2. Automatically creates a Bill of Entry (Draft) linked to that Invoice.
        """
        if not self.receipt_name:
            frappe.throw("No Purchase Receipt is linked to this Shipment Import.")

        pr_doc = frappe.get_doc("Purchase Receipt", self.receipt_name)
        if pr_doc.docstatus != 1:
            frappe.throw(
                f"Purchase Receipt <b>{self.receipt_name}</b> must be <b>Submitted</b> "
                f"before creating a Purchase Invoice. Current status: {pr_doc.docstatus}"
            )

        if not self.pi_naming_series:
            frappe.throw("Please select a Purchase Invoice Naming Series before creating the invoice.")

        if self.invoice_name:
            frappe.throw(
                f"A Purchase Invoice <b>{self.invoice_name}</b> already exists for this shipment."
            )

        # ── Step 1: Create Purchase Invoice from Purchase Receipt ────
        try:
            from erpnext.stock.doctype.purchase_receipt.purchase_receipt import make_purchase_invoice
            pi = make_purchase_invoice(self.receipt_name)
            pi.naming_series = self.pi_naming_series
            pi.insert(ignore_permissions=True)
            self.db_set("invoice_name", pi.name)
            frappe.db.commit()
        except Exception as exc:
            frappe.log_error(frappe.get_traceback(), "Shipment Import – Create PI Error")
            frappe.throw(f"Failed to create Purchase Invoice: {exc}")

        # ── Step 2: Auto-create Bill of Entry (Draft) ────────────────
        boe_name = None
        boe_error = None
        try:
            boe = frappe.new_doc("Bill of Entry")
            boe.purchase_invoice = pi.name
            boe.posting_date = nowdate()
            boe.company = pr_doc.company
            boe.supplier = pr_doc.supplier
            # Bill of Entry No will be filled by the user after BOE arrives
            boe.insert(ignore_permissions=True, ignore_mandatory=True)
            boe_name = boe.name
            self.db_set("bill_of_entry_name", boe_name)
            frappe.db.commit()
        except Exception as exc:
            frappe.log_error(frappe.get_traceback(), "Shipment Import – Create BOE Error")
            boe_error = str(exc)

        return {
            "invoice": pi.name,
            "bill_of_entry": boe_name,
            "boe_error": boe_error,
        }


# ─────────────────────────────────────────────
# Background worker functions
# ─────────────────────────────────────────────

def run_validation(docname):
    """
    Reads every data row from the Excel file and compares
    Qty & Rate against the matching Purchase Order line.

    Expected Excel columns (1-indexed):
      Col 1 – Item description (ignored)
      Col 2 – Item code (ignored)
      Col 3 – Qty (in Excel units; will be multiplied ×1000 to match PO)
      Col 4 – Rate (in Excel units; will be divided ÷1000 to match PO)
      Col 5 – PO Reference in format  <PO_NUMBER>-<LINE_IDX>  e.g. 12345-1
    """
    doc = frappe.get_doc("Shipment Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        file_path = file_doc.get_full_path()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        errors = []
        ok_rows = 0

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[4]:
                continue

            qty_exc = flt(row[2])
            rate_exc = flt(row[3])
            po_val_exc = str(row[4]).strip()

            # ── Parse PO reference ──────────────────────────────────
            # Look for optional Naming Series in Excel Column F (index 5)
            excel_series = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            
            try:
                po_parts = po_val_exc.split("-")
                base_po_num = po_parts[0]
                line_idx = int(po_parts[1])
                
                # Smart fallback for PO Name
                if excel_series:
                    po_name = f"{excel_series}{base_po_num}"
                elif doc.po_prefix:
                    po_name = f"{doc.po_prefix}{base_po_num}"
                else:
                    po_name = base_po_num # Fallback if full name is passed directly
                    
            except Exception:
                errors.append(
                    f"Row {row_idx} ❌  Cannot parse PO reference '{po_val_exc}'. "
                    f"Expected format: PO_NUMBER-LINE_INDEX  (e.g. 12345-1)"
                )
                continue

            # ── Check PO exists ─────────────────────────────────────
            if not frappe.db.exists("Purchase Order", po_name):
                errors.append(
                    f"Row {row_idx} ❌  Purchase Order '{po_name}' not found in ERPNext."
                )
                continue

            # ── Fetch PO line item ──────────────────────────────────
            po_item = frappe.db.get_value(
                "Purchase Order Item",
                {"parent": po_name, "idx": line_idx},
                ["qty", "rate", "item_code", "name"],
                as_dict=True,
            )
            if not po_item:
                errors.append(
                    f"Row {row_idx} ❌  Line {line_idx} not found in Purchase Order '{po_name}'."
                )
                continue

            row_ok = True

            # ── Quantity check ──────────────────────────────────────
            expected_qty = qty_exc * 1000
            if abs(expected_qty - po_item.qty) > 0.01:
                errors.append(
                    f"Row {row_idx} ❌  QTY MISMATCH — Item: {po_item.item_code} | "
                    f"PO: {po_name} Line {line_idx} | "
                    f"Excel Qty: {qty_exc} (×1000 = {expected_qty}) | "
                    f"PO Qty: {po_item.qty}"
                )
                row_ok = False

            # ── Rate check ──────────────────────────────────────────
            expected_rate = rate_exc / 1000
            if abs(expected_rate - po_item.rate) > 0.000001:
                errors.append(
                    f"Row {row_idx} ❌  RATE MISMATCH — Item: {po_item.item_code} | "
                    f"PO: {po_name} Line {line_idx} | "
                    f"Excel Rate: {rate_exc} (÷1000 = {expected_rate:.6f}) | "
                    f"PO Rate: {po_item.rate:.6f}"
                )
                row_ok = False

            if row_ok:
                ok_rows += 1

        # ── Build log and set status ────────────────────────────────
        if not errors:
            log_text = f"✅ All {ok_rows} row(s) validated successfully. No mismatches found."
            doc.db_set("status", "Validated")
        else:
            log_text = (
                f"❌ Validation Failed — {len(errors)} issue(s) found "
                f"({ok_rows} row(s) were OK):\n\n"
                + "\n".join(errors)
            )
            doc.db_set("status", "Failed")

        doc.db_set("validation_log", log_text)
        frappe.db.commit()

    except Exception:
        err = frappe.get_traceback()
        frappe.log_error(err, "Shipment Import – Validation Error")
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ System error during validation:\n\n{err}")
        frappe.db.commit()


def run_processing(docname):
    """
    Reads all validated rows from the Excel file and creates a SINGLE
    combined Purchase Receipt (saved as DRAFT) with line items from ALL Purchase Orders.
    Each line item still carries its individual PO reference for traceability.
    """
    doc = frappe.get_doc("Shipment Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        file_path = file_doc.get_full_path()
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active

        # ── Collect all rows in order ───────────────────────────────
        rows_to_process = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[4]:
                continue

            qty_exc = flt(row[2])
            rate_exc = flt(row[3])
            po_val_exc = str(row[4]).strip()

            # Look for optional Naming Series in Excel Column F (index 5)
            excel_series = str(row[5]).strip() if len(row) > 5 and row[5] else ""

            try:
                po_parts = po_val_exc.split("-")
                base_po_num = po_parts[0]
                line_idx = int(po_parts[1])
                
                # Smart fallback for PO Name
                if excel_series:
                    po_name = f"{excel_series}{base_po_num}"
                elif doc.po_prefix:
                    po_name = f"{doc.po_prefix}{base_po_num}"
                else:
                    po_name = base_po_num
            except Exception:
                continue

            rows_to_process.append(
                {
                    "po_name": po_name,
                    "line_idx": line_idx,
                    "qty": qty_exc * 1000,
                    "rate": rate_exc / 1000,
                    "row_idx": row_idx,
                }
            )

        if not rows_to_process:
            doc.db_set("status", "Failed")
            doc.db_set("receipts_log", "❌ No valid rows found to process.")
            frappe.db.commit()
            return

        # ── Determine company from first PO ─────────────────────────
        first_po = frappe.get_doc("Purchase Order", rows_to_process[0]["po_name"])
        company = first_po.company
        default_wh = frappe.db.get_single_value("Stock Settings", "default_warehouse")

        # ── Build one combined Purchase Receipt (DRAFT) ─────────────
        pr = frappe.new_doc("Purchase Receipt")
        pr.naming_series = doc.pr_naming_series
        pr.supplier = doc.supplier
        pr.posting_date = nowdate()
        pr.company = company

        added_lines = []
        errors = []
        po_set = set()

        for item_data in rows_to_process:
            po_name = item_data["po_name"]
            line_idx = item_data["line_idx"]

            po_item = frappe.db.get_value(
                "Purchase Order Item",
                {"parent": po_name, "idx": line_idx},
                [
                    "item_code", "item_name", "uom",
                    "warehouse", "name", "conversion_factor",
                ],
                as_dict=True,
            )

            if not po_item:
                errors.append(
                    f"❌  Skipped Row {item_data['row_idx']}: "
                    f"Line {line_idx} not found in PO {po_name}."
                )
                continue

            pr.append(
                "items",
                {
                    "item_code": po_item.item_code,
                    "item_name": po_item.item_name,
                    "qty": item_data["qty"],
                    "rate": item_data["rate"],
                    "uom": po_item.uom,
                    "warehouse": po_item.warehouse or default_wh,
                    "purchase_order": po_name,
                    "purchase_order_item": po_item.name,
                    "conversion_factor": po_item.conversion_factor or 1,
                },
            )

            added_lines.append(
                f"  • Row {item_data['row_idx']}: {po_item.item_code} | "
                f"PO: {po_name} Line {line_idx} | "
                f"Qty: {item_data['qty']} | Rate: {item_data['rate']:.4f}"
            )
            po_set.add(po_name)

        if not pr.items:
            doc.db_set("status", "Failed")
            doc.db_set("receipts_log", "❌ No items could be added to Purchase Receipt.\n\n" + "\n".join(errors))
            frappe.db.commit()
            return

        # ── Save as Draft ────────────────────────────────────────────
        try:
            pr.insert(ignore_permissions=True)
            # pr is saved as Draft (docstatus=0) — user must review and submit manually

            log_parts = [
                f"✅  Purchase Receipt <b>{pr.name}</b> created as DRAFT.",
                f"    ⚠️  Please review and Submit the receipt manually to post stock entries.",
                f"",
                f"    Supplier : {doc.supplier}",
                f"    Company  : {company}",
                f"    Date     : {nowdate()}",
                f"    POs covered ({len(po_set)}): {', '.join(sorted(po_set))}",
                f"    Total lines: {len(pr.items)}",
                "",
                "LINE DETAILS:",
                *added_lines,
            ]

            if errors:
                log_parts += ["", "SKIPPED LINES:", *errors]

            doc.db_set("status", "Completed")
            doc.db_set("receipt_name", pr.name)
            doc.db_set("receipts_log", "\n".join(log_parts))

        except Exception as exc:
            frappe.log_error(frappe.get_traceback(), "Shipment Import – PR Insert Error")
            doc.db_set("status", "Failed")
            doc.db_set(
                "receipts_log",
                f"❌ Failed to save Purchase Receipt:\n{exc}\n\nSkips:\n" + "\n".join(errors),
            )

        frappe.db.commit()

    except Exception:
        err = frappe.get_traceback()
        frappe.log_error(err, "Shipment Import – Processing Error")
        doc.db_set("status", "Failed")
        doc.db_set("receipts_log", f"❌ System error during processing:\n\n{err}")
        frappe.db.commit()
