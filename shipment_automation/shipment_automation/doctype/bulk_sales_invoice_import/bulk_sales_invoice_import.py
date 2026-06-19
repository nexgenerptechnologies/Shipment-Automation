import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date, validate_2_of_3


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Sales Invoice Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Sales Invoice Import"
    
    headers = [
        "Posting Date", "Sales Invoice Number (Optional)", "Delivery Note Number",
        "Sales Order Number", "Customer Name", "Item Code", "Item Name",
        "Description", "Quantity", "Rate", "Payment Due Date", "Update Stock (Yes/No)"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Sales_Invoice_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkSalesInvoiceImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_sales_invoice_import.bulk_sales_invoice_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("si_log", "⏳ Creating Sales Invoices. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_sales_invoice_import.bulk_sales_invoice_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "si_num": ["Sales Invoice Number (Optional)", "Sales Invoice Number", "SI Number"],
        "posting_date": ["Posting Date"],
        "payment_due_date": ["Payment Due Date", "Due Date"],
        "customer": ["Customer Name", "Customer"],
        "dn_num": ["Delivery Note Number", "DN Number"],
        "so_num": ["Sales Order Number", "SO Number"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "update_stock": ["Update Stock (Yes/No)", "Update Stock"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Sales Invoice Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        today = nowdate()
        
        requested_so_qty = {}
        requested_dn_qty = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            si_id_val = row[col_map["si_num"]] if col_map.get("si_num") is not None else None
            si_id = str(si_id_val).strip() if si_id_val is not None and str(si_id_val).strip() else ""
            customer = str(row[col_map["customer"]]).strip() if col_map.get("customer") is not None else ""
            dn_num = str(row[col_map["dn_num"]]).strip() if col_map.get("dn_num") is not None and row[col_map["dn_num"]] else ""
            so_num = str(row[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and row[col_map["so_num"]] else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None else ""
            
            posting_date = parse_excel_date(row[col_map["posting_date"]])

            if customer:
                customer_id = frappe.db.get_value("Customer", {"customer_name": customer}, "name")
                if not customer_id and not frappe.db.exists("Customer", customer):
                    row_errors.append(f"Customer \x27{customer}\x27 not found.")
            else:
                row_errors.append("Customer is mandatory.")
            
            if dn_num and not frappe.db.exists("Delivery Note", dn_num):
                row_errors.append(f"Delivery Note '{dn_num}' not found.")
            
            if so_num and not frappe.db.exists("Sales Order", so_num):
                row_errors.append(f"Sales Order '{so_num}' not found.")
            
            if not item_code or not frappe.db.exists("Item", item_code):
                row_errors.append(f"Item '{item_code}' not found.")
            else:
                # 2-of-3 column match
                it_doc = frappe.get_doc("Item", item_code)
                if not validate_2_of_3(it_doc, item_code, item_name, description):
                    row_errors.append("2-of-3 match failed (Code/Name/Description).")

            rate_exc = flt(row[col_map["rate"]]) if col_map.get("rate") is not None and row[col_map["rate"]] else 0
            qty_exc = flt(row[col_map["quantity"]]) if col_map.get("quantity") is not None and row[col_map["quantity"]] else 0
            
            if dn_num and item_code:
                key = (dn_num, item_code)
                requested_dn_qty[key] = requested_dn_qty.get(key, 0) + qty_exc
                dn_item = frappe.db.get_value("Delivery Note Item", {"parent": dn_num, "item_code": item_code}, ["rate", "qty"], as_dict=True)
                if dn_item:
                    if rate_exc and flt(dn_item.rate) != rate_exc:
                        row_errors.append(f"Rate {rate_exc} does not match Delivery Note {dn_num} rate ({dn_item.rate}).")
                    if requested_dn_qty[key] > flt(dn_item.qty):
                        row_errors.append(f"Cumulative quantity {requested_dn_qty[key]} exceeds Delivery Note {dn_num} quantity ({dn_item.qty}).")
            
            elif so_num and item_code:
                key = (so_num, item_code)
                requested_so_qty[key] = requested_so_qty.get(key, 0) + qty_exc
                so_item = frappe.db.get_value("Sales Order Item", {"parent": so_num, "item_code": item_code}, ["rate", "qty"], as_dict=True)
                if so_item:
                    if rate_exc and flt(so_item.rate) != rate_exc:
                        row_errors.append(f"Rate {rate_exc} does not match Sales Order {so_num} rate ({so_item.rate}).")
                    if requested_so_qty[key] > flt(so_item.qty):
                        row_errors.append(f"Cumulative quantity {requested_so_qty[key]} exceeds Sales Order {so_num} quantity ({so_item.qty}).")

            if posting_date and getdate(posting_date) > getdate(today):
                row_errors.append(f"Posting Date '{posting_date}' is a future date.")

            if si_id and frappe.db.exists("Sales Invoice", si_id):
                row_errors.append(f"Duplicate Sales Invoice Number '{si_id}' already exists.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Sales Invoice Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Group by SI ID if provided, else by Customer + Posting Date
        si_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            si_id_val = row[col_map["si_num"]] if col_map.get("si_num") is not None else None
            si_id = str(si_id_val).strip() if si_id_val is not None and str(si_id_val).strip() else ""
            cust = str(row[col_map["customer"]]).strip()
            date = parse_excel_date(row[col_map["posting_date"]]) or nowdate()
            
            group_key = si_id if si_id else (cust, date)
            if group_key not in si_groups: si_groups[group_key] = []
            si_groups[group_key].append(row)

        created = []
        for key, rows in si_groups.items():
            try:
                first = rows[0]
                customer_val = str(first[col_map["customer"]]).strip()
                customer = frappe.db.get_value("Customer", {"customer_name": customer_val}, "name")
                if not customer: customer = customer_val
                si_id_val = first[col_map["si_num"]] if col_map.get("si_num") is not None else None
                si_id = str(si_id_val).strip() if si_id_val is not None and str(si_id_val).strip() else ""
                posting_date = parse_excel_date(first[col_map["posting_date"]]) or nowdate()
                update_stock = 1 if str(first[col_map["update_stock"]]).strip().lower() in ["yes", "y", "1"] else 0
                
                si = frappe.new_doc("Sales Invoice")
                if si_id:
                    si.name = si_id
                    si.flags.ignore_autoname = True
                    si.naming_series = ""
                
                si.customer = customer
                si.set_posting_time = 1
                si.posting_date = posting_date
                due_date_val = parse_excel_date(first[col_map["payment_due_date"]]) if col_map.get("payment_due_date") is not None else None
                if due_date_val:
                    si.due_date = due_date_val
                si.update_stock = update_stock
                
                for r in rows:
                    dn_num = str(r[col_map["dn_num"]]).strip() if col_map.get("dn_num") is not None and r[col_map["dn_num"]] else ""
                    so_num = str(r[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and r[col_map["so_num"]] else ""
                    item_code = str(r[col_map["item_code"]]).strip()
                    qty = flt(r[col_map["quantity"]])
                    rate = flt(r[col_map["rate"]])
                    
                    item_row = si.append("items", {
                        "item_code": item_code,
                        "qty": qty,
                        "rate": rate,
                        "delivery_note": dn_num,
                        "sales_order": so_num
                    })
                    
                    # Logic to pull from DN or SO if provided
                    if dn_num:
                        dn_item = frappe.db.get_value("Delivery Note Item", 
                            {"parent": dn_num, "item_code": item_code}, 
                            ["name", "rate", "warehouse"], as_dict=True)
                        if dn_item:
                            item_row.dn_detail = dn_item.name
                            item_row.warehouse = dn_item.warehouse
                            if not rate: item_row.rate = dn_item.rate
                    
                    elif so_num:
                        so_item = frappe.db.get_value("Sales Order Item", 
                            {"parent": so_num, "item_code": item_code}, 
                            ["name", "rate"], as_dict=True)
                        if so_item:
                            item_row.so_detail = so_item.name
                            if not rate: item_row.rate = so_item.rate

                si.flags.ignore_permissions = True
                
                frappe.db.savepoint("si_create")
                try:
                    si.set_missing_values()
                    si.insert(ignore_permissions=True)
                    
                    if si_id and si.name != si_id:
                        frappe.rename_doc("Sales Invoice", si.name, si_id, force=True, ignore_permissions=True)
                        si.name = si_id
                    
                    # si.submit() # Optional
                    created.append(f"✅ {si.name}")
                except Exception as e:
                    frappe.db.rollback(save_point="si_create")
                    created.append(f"❌ Error creating SI {si_id or customer}: {str(e)}")
            except Exception as e:
                frappe.db.rollback()
                created.append(f"❌ Error creating SI: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("si_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("si_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()

