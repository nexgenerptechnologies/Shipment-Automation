import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate
import openpyxl
from io import BytesIO


@frappe.whitelist()
def download_template():
    """Generates and downloads the PO Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Import Template"
    
    headers = [
        "Supplier", "Purchase Order Number", "Date", "Required By",
        "Item Code", "Item Name", "Description", "Quantity",
        "Rate", "Item Group", "HSN/SAC", "Line Number"
    ]
    ws.append(headers)
    
    # Simple formatting for headers
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Shipment_PO_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class ShipmentPOImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        if not self.po_naming_series:
            frappe.throw("Please select a PO Naming Series before uploading.")
        if not self.item_naming_series:
            frappe.throw("Please select an Item Naming Series before uploading.")
        self.db_set("status", "Validating")
        self.db_set("creation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.shipment_po_import.shipment_po_import.run_po_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the Excel first.")
        self.db_set("status", "Processing")
        self.db_set("creation_log", "⏳ Creating Purchase Orders. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.shipment_po_import.shipment_po_import.run_po_creation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Purchase Order creation started. Refresh in a few seconds."


def find_duplicate_item(e_code, e_name, e_desc):
    """Checks if an item matching 2 out of 3 fields exists."""
    if e_code and e_name:
        res = frappe.db.get_value("Item", {"name": e_code, "item_name": e_name}, "name")
        if res: return res
    if e_code and e_desc:
        res = frappe.db.get_value("Item", {"name": e_code, "description": e_desc}, "name")
        if res: return res
    if e_name and e_desc:
        res = frappe.db.get_value("Item", {"item_name": e_name, "description": e_desc}, "name")
        if res: return res
    return None


def get_column_map(sheet):
    """Maps header names to column indices."""
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected_headers = {
        "supplier": ["Supplier", "Supplier Name"],
        "po_num": ["Purchase Order Number", "PO Number", "PO #"],
        "date": ["Date", "Posting Date"],
        "required_by": ["Required By", "Delivery Date"],
        "item_code": ["Item Code", "Item"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate", "Price"],
        "item_group": ["Item Group"],
        "hsn_sac": ["HSN/SAC", "HSN Code", "SAC Code"],
        "line_number": ["Line Number", "Line #"]
    }
    
    for idx, cell_value in enumerate(header_row):
        if not cell_value: continue
        clean_val = str(cell_value).strip().lower()
        for key, aliases in expected_headers.items():
            if any(alias.lower() == clean_val for alias in aliases):
                mapping[key] = idx
    return mapping


def run_po_validation(docname):
    doc = frappe.get_doc("Shipment PO Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.po_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active

        col_map = get_column_map(sheet)
        required_cols = ["item_code", "quantity", "rate", "po_num"]
        missing_cols = [c for c in required_cols if c not in col_map]
        
        if missing_cols:
            frappe.throw(f"Missing required columns in Excel: {', '.join(missing_cols)}")

        errors = []
        ok_rows = 0
        items_to_create = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue

            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None and row[col_map["item_code"]] else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None and row[col_map["item_name"]] else ""
            desc      = str(row[col_map["description"]]).strip() if col_map.get("description") is not None and row[col_map["description"]] else ""
            qty       = flt(row[col_map["quantity"]]) if col_map.get("quantity") is not None else 0
            rate      = flt(row[col_map["rate"]]) if col_map.get("rate") is not None else 0
            supplier  = str(row[col_map["supplier"]]).strip() if col_map.get("supplier") is not None and row[col_map["supplier"]] else ""
            po_num    = str(row[col_map["po_num"]]).strip() if col_map.get("po_num") is not None and row[col_map["po_num"]] else ""
            item_group = str(row[col_map["item_group"]]).strip() if col_map.get("item_group") is not None and row[col_map["item_group"]] else ""
            hsn_code   = str(row[col_map["hsn_sac"]]).strip() if col_map.get("hsn_sac") is not None and row[col_map["hsn_sac"]] else ""

            if not po_num:
                errors.append(f"Row {row_idx} ❌ PO Number is missing.")
                continue
                
            row_ok = True
            if supplier and not frappe.db.exists("Supplier", supplier):
                errors.append(f"Row {row_idx} ❌ Supplier '{supplier}' not found.")
                row_ok = False

            if qty <= 0:
                errors.append(f"Row {row_idx} ❌ Quantity must be > 0.")
                row_ok = False

            if not item_code:
                errors.append(f"Row {row_idx} ❌ Item Code is empty.")
                row_ok = False
            else:
                if not frappe.db.exists("Item", item_code):
                    duplicate = find_duplicate_item(item_code, item_name, desc)
                    if duplicate:
                        errors.append(f"Row {row_idx} ❌ Duplicate Found: Item '{duplicate}'. Update Excel Item Code to match.")
                        row_ok = False
                    else:
                        if not item_group or not hsn_code:
                            errors.append(f"Row {row_idx} ❌ New Item '{item_code}' requires Item Group and HSN/SAC.")
                            row_ok = False
                        else:
                            if not frappe.db.exists("Item Group", item_group):
                                errors.append(f"Row {row_idx} ❌ Item Group '{item_group}' not found in system.")
                                row_ok = False
                            
                            if not any(i['item_code'] == item_code for i in items_to_create):
                                items_to_create.append({"item_code": item_code, "item_name": item_name or item_code, "description": desc, "item_group": item_group, "hsn_code": hsn_code, "uom": "Nos"})

            if row_ok: ok_rows += 1

        if not errors:
            log = [f"✅ {ok_rows} row(s) validated."]
            if items_to_create:
                log.append(f"\n✨ {len(items_to_create)} New Items will be created:")
                for it in items_to_create: log.append(f"  • {it['item_code']} | {it['item_name']}")
            doc.db_set("status", "Validated")
            doc.db_set("creation_log", "\n".join(log))
        else:
            doc.db_set("status", "Draft")
            doc.db_set("creation_log", f"❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Shipment PO Import – Validation Error")
        doc.db_set("status", "Failed")
        doc.db_set("creation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_po_creation(docname):
    doc = frappe.get_doc("Shipment PO Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.po_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)

        item_mapping = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None and row[col_map["item_code"]] else ""
            if item_code and not frappe.db.exists("Item", item_code):
                if item_code not in item_mapping:
                    new_item = frappe.new_doc("Item")
                    new_item.naming_series = doc.item_naming_series
                    new_item.item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None and row[col_map["item_name"]] else item_code
                    new_item.description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None and row[col_map["description"]] else ""
                    new_item.item_group = str(row[col_map["item_group"]]).strip() if col_map.get("item_group") is not None and row[col_map["item_group"]] else ""
                    new_item.gst_hsn_code = str(row[col_map["hsn_sac"]]).strip() if col_map.get("hsn_sac") is not None and row[col_map["hsn_sac"]] else ""
                    new_item.stock_uom = "Nos"
                    new_item.insert(ignore_permissions=True)
                    item_mapping[item_code] = new_item.name

        po_map = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row) or not row[col_map["po_num"]]: continue
            po_num = str(row[col_map["po_num"]]).strip()
            po_map.setdefault(po_num, {"supplier": str(row[col_map["supplier"]]).strip() if col_map.get("supplier") is not None else doc.supplier, "items": []})
            po_map[po_num]["items"].append({
                "item_code": item_mapping.get(str(row[col_map["item_code"]]).strip(), str(row[col_map["item_code"]]).strip()),
                "qty": flt(row[col_map["quantity"]]),
                "rate": flt(row[col_map["rate"]]),
                "line_number": str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""
            })

        created = []
        company = frappe.db.get_single_value("Global Defaults", "default_company")
        for p_num, data in po_map.items():
            po = frappe.new_doc("Purchase Order")
            po.naming_series, po.supplier, po.company = doc.po_naming_series, data["supplier"], company
            for item in data["items"]:
                po.append("items", {"item_code": item["item_code"], "qty": item["qty"], "rate": item["rate"], "schedule_date": nowdate()})
            po.insert(ignore_permissions=True)
            
            import re
            match = re.search(r'(\d+)$', p_num)
            base_number = match.group(1) if match else p_num
            for idx, item in enumerate(po.items, start=1):
                item.db_set("line_number", data["items"][idx-1]["line_number"] or f"{base_number}-{idx}")
            created.append(f"✅ {po.name} (Excel: {p_num})")

        doc.db_set("status", "Completed")
        doc.db_set("creation_log", "CREATED:\n" + "\n".join(created))
        frappe.db.commit()

    except Exception:
        frappe.log_error(frappe.get_traceback(), "Shipment PO Import – Creation Error")
        doc.db_set("status", "Failed")
        doc.db_set("creation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
