import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, nowdate
import openpyxl
from io import BytesIO
import datetime


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Payment Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Payment Import Template"
    
    headers = [
        "Payment ID (Link rows)", "Payment Type", "Posting Date", "Mode of Payment",
        "Company Bank Account", "Party Type", "Party", "Total Amount", 
        "Reference Type", "Reference Name", "Allocated Amount", "UTR/Ref No", "Ref Date"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Payment_Entry_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkPaymentEntryImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_payment_entry_import.bulk_payment_entry_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("payment_log", "⏳ Creating Payment Entries. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_payment_entry_import.bulk_payment_entry_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "p_id": ["Payment ID (Link rows)", "Payment ID"],
        "p_type": ["Payment Type"],
        "posting_date": ["Posting Date"],
        "mop": ["Mode of Payment"],
        "company_bank": ["Company Bank Account", "Company Account", "Account Paid To"],
        "party_type": ["Party Type"],
        "party": ["Party"],
        "total_amount": ["Total Amount", "Amount", "Paid Amount"],
        "ref_type": ["Reference Type"],
        "ref_name": ["Reference Name", "Invoice No"],
        "allocated_amount": ["Allocated Amount"],
        "utr_no": ["UTR/Ref No", "Reference No", "Cheque No"],
        "ref_date": ["Ref Date", "Reference Date", "Cheque Date"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def parse_excel_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime("%Y-%m-%d")
    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"]:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    try:
        res = getdate(date_val)
        return res.strftime("%Y-%m-%d") if res else None
    except:
        return None


def resolve_party_id(party_type, party_name):
    if not party_type or not party_name:
        return party_name
    party_field = frappe.scrub(party_type) + "_name"
    party_id = frappe.db.get_value(party_type, {party_field: party_name}, "name")
    if not party_id:
        party_id = frappe.db.get_value(party_type, {"name": party_name}, "name")
    return party_id or party_name

def run_validation(docname):
    doc = frappe.get_doc("Bulk Payment Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        
        missing_cols = []
        for req in ["p_type", "posting_date", "mop", "party_type", "party", "total_amount"]:
            if req not in col_map: missing_cols.append(req)
        
        if missing_cols:
            errors.append(f"❌ Missing required columns in Excel header: {', '.join(missing_cols)}")
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\\n\\n" + "\\n".join(errors))
            frappe.db.commit()
            return
            
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            p_type = str(row[col_map["p_type"]]).strip() if col_map.get("p_type") is not None and row[col_map["p_type"]] else ""
            p_type = p_type.capitalize()
            if p_type not in ["Receive", "Pay", "Internal Transfer"]:
                row_errors.append(f"Invalid Payment Type '{p_type}'. Must be Receive, Pay, or Internal Transfer.")
                
            mop = str(row[col_map["mop"]]).strip() if col_map.get("mop") is not None and row[col_map["mop"]] else ""
            if mop and not frappe.db.exists("Mode of Payment", mop):
                row_errors.append(f"Mode of Payment '{mop}' not found.")
                
            company_bank = str(row[col_map["company_bank"]]).strip() if col_map.get("company_bank") is not None and row[col_map["company_bank"]] else ""
            if company_bank and not frappe.db.exists("Account", company_bank):
                row_errors.append(f"Company Bank Account '{company_bank}' not found.")
                
            ptype = str(row[col_map["party_type"]]).strip() if col_map.get("party_type") is not None and row[col_map["party_type"]] else ""
            pname = str(row[col_map["party"]]).strip() if col_map.get("party") is not None and row[col_map["party"]] else ""
            
            if ptype and pname:
                party_id = resolve_party_id(ptype, pname)
                if not frappe.db.exists(ptype, party_id):
                    row_errors.append(f"Party '{pname}' of type '{ptype}' not found.")
                    
            ref_type = str(row[col_map["ref_type"]]).strip() if col_map.get("ref_type") is not None and row[col_map["ref_type"]] else ""
            ref_name = str(row[col_map["ref_name"]]).strip() if col_map.get("ref_name") is not None and row[col_map["ref_name"]] else ""
            if ref_type and ref_name:
                if not frappe.db.exists(ref_type, ref_name):
                    row_errors.append(f"{ref_type} '{ref_name}' not found.")
                    
            posting_date = row[col_map["posting_date"]]
            parsed_pd = parse_excel_date(posting_date)
            if not parsed_pd:
                row_errors.append("Invalid Posting Date.")
                
            total_amt = flt(row[col_map["total_amount"]]) if col_map.get("total_amount") is not None else 0.0
            if total_amt <= 0:
                row_errors.append("Total Amount must be greater than 0.")
                
            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
                
        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", "✅ All entries validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\\n\\n" + "\\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Payment Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Grouping
        groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            p_id_val = str(row[col_map["p_id"]]).strip() if col_map.get("p_id") is not None and row[col_map["p_id"]] else ""
            p_type = str(row[col_map["p_type"]]).strip().capitalize() if col_map.get("p_type") is not None else ""
            pname = str(row[col_map["party"]]).strip() if col_map.get("party") is not None and row[col_map["party"]] else ""
            total_amt = flt(row[col_map["total_amount"]]) if col_map.get("total_amount") is not None else 0.0
            posting_date = parse_excel_date(row[col_map["posting_date"]]) if col_map.get("posting_date") is not None else ""
            
            if p_id_val:
                p_id = p_id_val
            else:
                p_id = f"{p_type}_{posting_date}_{pname}_{total_amt}"
            
            if p_id not in groups: groups[p_id] = []
            groups[p_id].append(row)

        created = []
        for p_id, rows in groups.items():
            try:
                first = rows[0]
                pe = frappe.new_doc("Payment Entry")
                
                pe.payment_type = str(first[col_map["p_type"]]).strip().capitalize()
                pe.posting_date = parse_excel_date(first[col_map["posting_date"]])
                pe.mode_of_payment = str(first[col_map["mop"]]).strip() if first[col_map["mop"]] else ""
                
                ptype = str(first[col_map["party_type"]]).strip() if col_map.get("party_type") is not None and first[col_map["party_type"]] else ""
                pname = str(first[col_map["party"]]).strip() if col_map.get("party") is not None and first[col_map["party"]] else ""
                if ptype and pname:
                    pe.party_type = ptype
                    pe.party = resolve_party_id(ptype, pname)
                
                total_amt = flt(first[col_map["total_amount"]])
                pe.paid_amount = total_amt
                pe.received_amount = total_amt
                
                company_bank = str(first[col_map["company_bank"]]).strip() if col_map.get("company_bank") is not None and first[col_map["company_bank"]] else ""
                
                # Fetch default party account if possible
                party_account = ""
                if pe.party_type and pe.party:
                    party_account = frappe.db.get_value("Party Account", {"parent": pe.party, "parenttype": pe.party_type}, "account")
                    if not party_account:
                        party_account = frappe.db.get_value(pe.party_type, pe.party, "default_account")
                        
                # Set up accounts
                if pe.payment_type == "Receive":
                    if party_account: pe.paid_from = party_account
                    if company_bank: pe.paid_to = company_bank
                elif pe.payment_type == "Pay":
                    if party_account: pe.paid_to = party_account
                    if company_bank: pe.paid_from = company_bank
                
                utr = str(first[col_map["utr_no"]]).strip() if col_map.get("utr_no") is not None and first[col_map["utr_no"]] else ""
                if utr:
                    pe.reference_no = utr
                    ref_date = first[col_map["ref_date"]] if col_map.get("ref_date") is not None else None
                    pe.reference_date = parse_excel_date(ref_date) if ref_date else pe.posting_date

                # Let Frappe auto-set missing default accounts
                pe.setup_party_account_field()
                pe.set_missing_values()

                # Set References
                for r in rows:
                    ref_type = str(r[col_map["ref_type"]]).strip() if col_map.get("ref_type") is not None and r[col_map["ref_type"]] else ""
                    ref_name = str(r[col_map["ref_name"]]).strip() if col_map.get("ref_name") is not None and r[col_map["ref_name"]] else ""
                    alloc_amt = flt(r[col_map["allocated_amount"]]) if col_map.get("allocated_amount") is not None else 0.0
                    
                    if ref_type and ref_name and alloc_amt > 0:
                        pe.append("references", {
                            "reference_doctype": ref_type,
                            "reference_name": ref_name,
                            "allocated_amount": alloc_amt
                        })
                
                pe.flags.ignore_permissions = True
                pe.insert()
                created.append(f"✅ {pe.name}")
            except Exception as e:
                msg = str(e)
                if hasattr(e, 'message'): msg = e.message
                elif hasattr(e, 'args') and e.args: msg = e.args[0]
                from frappe.utils import strip_html
                msg = strip_html(str(msg))
                created.append(f"❌ {p_id}: {msg}")

        status = "Completed" if not any("❌" in log for log in created) else "Failed"
        doc.db_set("status", status)
        doc.db_set("payment_log", "SUMMARY:\\n" + "\\n".join(created))
        frappe.db.commit()
    except Exception as e:
        doc.db_set("status", "Failed")
        doc.db_set("payment_log", f"❌ Critical Error:\\n{str(e)}")
        frappe.db.commit()
