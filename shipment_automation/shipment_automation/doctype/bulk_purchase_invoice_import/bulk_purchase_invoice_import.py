import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Purchase Invoice & BOE Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk PI and BOE Import"
    
    headers = [
        "Purchase Receipt Number", "Purchase Receipt Date", "Supplier Name", 
        "Purchase Order Number", "Item Code", "Item Name", 
        "Description", "Quantity", "Rate", "Line Number",
        "Purchase Invoice Posting Date", "Purchase Invoice Number", "Purchase Invoice Date",
        "Bill of Entry Posting Date", "Bill of Entry Number", "Bill of Entry Date"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Invoice_BOE_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkPurchaseInvoiceImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_purchase_invoice_import.bulk_purchase_invoice_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_processing(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Invoices & BOEs. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_purchase_invoice_import.bulk_purchase_invoice_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "pr_num": ["Purchase Receipt Number", "PR Number"],
        "supplier": ["Supplier Name", "Supplier"],
        "po_num": ["Purchase Order Number", "PO Number"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "line_number": ["Line Number", "Line #"],
        "pi_post_date": ["Purchase Invoice Posting Date"],
        "pi_num": ["Purchase Invoice Number", "PI Number", "Invoice No"],
        "pi_date": ["Purchase Invoice Date", "PI Date", "Invoice Date"],
        "boe_post_date": ["Bill of Entry Posting Date"],
        "boe_num": ["Bill of Entry Number", "BOE Number", "BOE No"],
        "boe_date": ["Bill of Entry Date", "BOE Date"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def find_po_item_by_line(po_num, item_code, line_val):
    """STRICTLY finds a PO Item by Line Number."""
    meta = frappe.get_meta("Purchase Order Item")
    af = [f.fieldname for f in meta.fields]
    if "line_number" in af:
        res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "line_number": line_val, "item_code": item_code}, "name")
        if res: return res
    if "custom_line_number" in af:
        res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "custom_line_number": line_val, "item_code": item_code}, "name")
        if res: return res
    import re
    match = re.search(r'(\d+)$', po_num)
    if match:
        suffix = match.group(1)
        if line_val.startswith(f"{suffix}-"):
            try:
                idx_part = int(line_val.split("-")[-1])
                res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "idx": idx_part, "item_code": item_code}, "name")
                if res: return res
            except: pass
    return None


def parse_excel_date(date_val):
    if not date_val: return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        if date_val.day <= 12:
             try: return datetime.date(date_val.year, date_val.day, date_val.month).strftime("%Y-%m-%d")
             except ValueError: return date_val.strftime("%Y-%m-%d")
        return date_val.strftime("%Y-%m-%d")
    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError: continue
    try:
        res = getdate(date_val)
        return res.strftime("%Y-%m-%d") if res else None
    except: return None


def run_validation(docname):
    doc = frappe.get_doc("Bulk Purchase Invoice Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        today = nowdate()
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            
            pr_id = str(row[col_map["pr_num"]]).strip() if col_map.get("pr_num") is not None and row[col_map["pr_num"]] else ""
            po_num = str(row[col_map["po_num"]]).strip() if col_map.get("po_num") is not None and row[col_map["po_num"]] else ""
            supplier_name = str(row[col_map["supplier"]]).strip() if col_map.get("supplier") is not None else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None else ""
            qty_exc = flt(row[col_map["quantity"]]) if col_map.get("quantity") is not None else 0
            rate_exc = flt(row[col_map["rate"]]) if col_map.get("rate") is not None else 0
            line_val = str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""
            
            pi_id = str(row[col_map["pi_num"]]).strip() if col_map.get("pi_num") is not None and row[col_map["pi_num"]] else ""
            pi_post_date = parse_excel_date(row[col_map["pi_post_date"]]) if col_map.get("pi_post_date") is not None else None
            boe_id = str(row[col_map["boe_num"]]).strip() if col_map.get("boe_num") is not None and row[col_map["boe_num"]] else ""

            # ── Scenario Logic & PO Validation ──
            if pr_id:
                # SCENARIO 1: PR-to-PI
                if not frappe.db.exists("Purchase Receipt", pr_id):
                    row_errors.append(f"Purchase Receipt '{pr_id}' not found.")
                if not po_num:
                    row_errors.append("PO Number is mandatory when PR Number is provided.")
            elif po_num:
                # SCENARIO 2: PO-to-PI
                if not frappe.db.exists("Purchase Order", po_num):
                    row_errors.append(f"Purchase Order '{po_num}' not found.")
            # Scenario 3: Direct PI (No PR, No PO) - Allowed as fallback

            # ── Strict PO Matching (if PO exists) ──
            if po_num and frappe.db.exists("Purchase Order", po_num):
                po_item_name = find_po_item_by_line(po_num, item_code, line_val) if line_val else frappe.db.get_value("Purchase Order Item", {"parent": po_num, "item_code": item_code}, "name")
                if not po_item_name:
                    row_errors.append(f"Item/Line '{line_val or item_code}' not found in PO '{po_num}'.")
                else:
                    pi_item = frappe.get_doc("Purchase Order Item", po_item_name)
                    # 7-decimal Rate Check
                    if abs(flt(pi_item.rate) - flt(rate_exc)) > 0.0000001:
                        row_errors.append(f"Rate mismatch: Excel {rate_exc} vs PO {pi_item.rate}")
                    # 2-of-3 column match
                    score = 0
                    if pi_item.item_code == item_code: score += 1
                    if pi_item.item_name == item_name: score += 1
                    if pi_item.description == description: score += 1
                    if score < 2:
                        row_errors.append("2-of-3 column match failed (Code/Name/Description).")

            # ── General Validations ──
            if not pi_id:
                row_errors.append("Purchase Invoice Number missing.")
            elif frappe.db.exists("Purchase Invoice", pi_id):
                row_errors.append(f"Duplicate Error: Invoice '{pi_id}' already exists.")

            if pi_post_date and getdate(pi_post_date) > getdate(today):
                row_errors.append(f"Invoice Posting Date '{pi_post_date}' is a future date.")

            # Overseas BOE Check
            if boe_id and supplier_name:
                gst_cat = frappe.db.get_value("Supplier", supplier_name, "gst_category")
                if gst_cat != "Overseas":
                    row_errors.append(f"Bill of Entry creation failed: Supplier '{supplier_name}' is '{gst_cat}', BOE only allowed for 'Overseas'.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Purchase Invoice Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        pi_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            pi_id = str(row[col_map["pi_num"]]).strip()
            if pi_id not in pi_groups: pi_groups[pi_id] = []
            pi_groups[pi_id].append(row)

        created = []
        for pi_id, rows in pi_groups.items():
            try:
                first = rows[0]
                pr_id = str(first[col_map["pr_num"]]).strip() if col_map.get("pr_num") is not None and first[col_map["pr_num"]] else ""
                po_num = str(first[col_map["po_num"]]).strip() if col_map.get("po_num") is not None and first[col_map["po_num"]] else ""
                supplier = str(first[col_map["supplier"]]).strip()
                pi_post_date = parse_excel_date(first[col_map["pi_post_date"]])
                pi_date = parse_excel_date(first[col_map["pi_date"]])
                boe_id = str(first[col_map["boe_num"]]).strip() if col_map.get("boe_num") is not None and first[col_map["boe_num"]] else ""
                boe_post_date = parse_excel_date(first[col_map["boe_post_date"]])
                boe_date = parse_excel_date(first[col_map["boe_date"]])

                if frappe.db.exists("Purchase Invoice", pi_id): continue

                # ── Create Purchase Invoice ──
                if pr_id:
                    # Scenario 1: PR-to-PI
                    from erpnext.buying.doctype.purchase_receipt.purchase_receipt import make_purchase_invoice
                    pi = make_purchase_invoice(pr_id)
                elif po_num:
                    # Scenario 2: PO-to-PI
                    from erpnext.buying.doctype.purchase_order.purchase_order import make_purchase_invoice
                    pi = make_purchase_invoice(po_num)
                    pi.update_stock = 1
                else:
                    # Scenario 3: Standalone
                    pi = frappe.new_doc("Purchase Invoice")
                    pi.update_stock = 1

                # Common Headers
                pi.name = pi_id
                pi.supplier = supplier
                pi.posting_date = pi_post_date or pi_date or nowdate()
                pi.bill_no = pi_id
                pi.bill_date = pi_date or pi_post_date or nowdate()
                
                # Naming Series
                pi_series = frappe.get_meta("Purchase Invoice").get_field("naming_series").options.split("\n")
                pi.naming_series = pi_series[0]
                for s in pi_series:
                    pfx = s.replace(".####", "").replace(".YY.", "").replace(".YYYY.", "").strip()
                    if pfx and pi_id.startswith(pfx):
                        pi.naming_series = s
                        break

                # Update Items (Handle partials or Standalone lines)
                if not pr_id and not po_num: # Only for Standalone
                    pi.items = []
                    for r in rows:
                        pi.append("items", {
                            "item_code": str(r[col_map["item_code"]]).strip(),
                            "qty": flt(r[col_map["quantity"]]),
                            "rate": flt(r[col_map["rate"]])
                        })
                
                pi.flags.ignore_permissions = True
                pi.db_insert()
                for i in pi.get("items"): i.db_insert()
                for t in pi.get("taxes"): t.db_insert()
                pi.run_method("on_update")
                pi.submit()

                # ── Create BOE (Overseas only) ──
                if boe_id and frappe.db.get_value("Supplier", supplier, "gst_category") == "Overseas":
                    boe = frappe.new_doc("Bill of Entry")
                    boe.name = boe_id
                    boe.purchase_invoice = pi.name
                    boe.posting_date = boe_post_date or boe_date or nowdate()
                    boe.bill_of_entry_number = boe_id
                    boe.bill_of_entry_date = boe_date or boe_post_date or nowdate()
                    boe.company = pi.company
                    boe.supplier = pi.supplier
                    boe.flags.ignore_permissions = True
                    boe.insert(ignore_mandatory=True)
                    created.append(f"✅ {pi_id} (BOE {boe_id})")
                else:
                    created.append(f"✅ {pi_id}")

            except Exception as e:
                created.append(f"❌ {pi_id}: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("processing_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
