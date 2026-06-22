import frappe
import openpyxl
from frappe.model.document import Document
from frappe.utils import flt, get_site_path
from io import BytesIO

class BulkBOMImport(Document):
    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validating BOM structure...")
        frappe.db.commit()
        frappe.enqueue("shipment_automation.shipment_automation.doctype.bulk_bom_import.bulk_bom_import.run_validation", docname=self.name, queue="long", timeout=3600)

    @frappe.whitelist()
    def start_processing(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Bill of Materials records. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue("shipment_automation.shipment_automation.doctype.bulk_bom_import.bulk_bom_import.run_processing", docname=self.name, queue="long", timeout=3600)

    @frappe.whitelist()
    def download_template(self):
        wb = openpyxl.Workbook()
        
        # Sheet 1: Materials
        ws1 = wb.active
        ws1.title = "BOM Materials"
        ws1.append([
            "Item to Manufacture (Item Code)", "Qty to Manufacture", 
            "Item Consume", "Qty Required"
        ])
        
        # Sheet 2: Operations
        ws2 = wb.create_sheet(title="BOM Operations")
        ws2.append([
            "Item to Manufacture (Item Code)", "Operation", 
            "Workstation", "Operation Time (Mins)", "Hour Rate", "Is Subcontracted"
        ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        frappe.response['filename'] = "Bulk_BOM_Import_Template.xlsx"
        frappe.response['filecontent'] = output.getvalue()
        frappe.response['type'] = 'binary'

def clean_val(val):
    if val is None or str(val).lower() == "none":
        return ""
    return str(val).strip()

def get_col_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        if "manufacture" in clean or "parent" in clean: mapping["parent"] = idx
        elif "qty to manufacture" in clean: mapping["qty"] = idx
        elif "consume" in clean or "component" in clean: mapping["child"] = idx
        elif "qty required" in clean or "child qty" in clean: mapping["child_qty"] = idx
        elif "operation" in clean and "time" not in clean: mapping["operation"] = idx
        elif "workstation" in clean and "type" not in clean: mapping["workstation"] = idx
        elif "operation time" in clean: mapping["op_time"] = idx
        elif "hour rate" in clean: mapping["hour_rate"] = idx
        elif "subcontracted" in clean: mapping["is_subcontracted"] = idx
    return mapping

def run_validation(docname):
    doc = frappe.get_doc("Bulk BOM Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.bom_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        
        errors = []
        
        # Validate Materials
        if "BOM Materials" in wb.sheetnames:
            ws1 = wb["BOM Materials"]
            col_map1 = get_col_map(ws1)
            for i, row in enumerate(ws1.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                parent = clean_val(row[col_map1.get("parent", -1)]) if "parent" in col_map1 else ""
                child = clean_val(row[col_map1.get("child", -1)]) if "child" in col_map1 else ""
                
                if not parent:
                    errors.append(f"Materials Row {i}: Item to Manufacture is missing.")
                elif not frappe.db.exists("Item", parent):
                    errors.append(f"Materials Row {i}: Item to Manufacture '{parent}' does not exist.")
                
                if child and not frappe.db.exists("Item", child):
                    errors.append(f"Materials Row {i}: Item Consume '{child}' does not exist.")
        
        # Validate Operations
        if "BOM Operations" in wb.sheetnames:
            ws2 = wb["BOM Operations"]
            col_map2 = get_col_map(ws2)
            for i, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): continue
                parent = clean_val(row[col_map2.get("parent", -1)]) if "parent" in col_map2 else ""
                op_name = clean_val(row[col_map2.get("operation", -1)]) if "operation" in col_map2 else ""
                workstation = clean_val(row[col_map2.get("workstation", -1)]) if "workstation" in col_map2 else ""
                
                if not parent:
                    errors.append(f"Operations Row {i}: Item to Manufacture is missing.")
                elif not frappe.db.exists("Item", parent):
                    errors.append(f"Operations Row {i}: Item to Manufacture '{parent}' does not exist.")
                
                if op_name and workstation and not frappe.db.exists("Workstation", workstation):
                    errors.append(f"Operations Row {i}: Workstation '{workstation}' does not exist.")

        if errors:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Validation Errors:\n" + "\n".join(errors[:50]) + ("\n...and more" if len(errors)>50 else ""))
        else:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", "✅ Validation Successful.")
        frappe.db.commit()
        
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()

def run_processing(docname):
    doc = frappe.get_doc("Bulk BOM Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.bom_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        
        boms = {}
        dependencies = {} # parent -> set of child items
        
        # Parse Materials
        if "BOM Materials" in wb.sheetnames:
            ws1 = wb["BOM Materials"]
            col_map1 = get_col_map(ws1)
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                parent = clean_val(row[col_map1.get("parent", -1)]) if "parent" in col_map1 else ""
                if not parent: continue
                
                if parent not in boms:
                    boms[parent] = {"qty": 1.0, "items": [], "operations": []}
                    dependencies[parent] = set()
                
                qty = flt(row[col_map1.get("qty", -1)]) if "qty" in col_map1 else 0
                if qty > 0: boms[parent]["qty"] = qty
                
                child = clean_val(row[col_map1.get("child", -1)]) if "child" in col_map1 else ""
                if child:
                    dependencies[parent].add(child)
                    boms[parent]["items"].append({
                        "item_code": child,
                        "qty": flt(row[col_map1.get("child_qty", -1)]) or 1.0
                    })

        # Parse Operations
        if "BOM Operations" in wb.sheetnames:
            ws2 = wb["BOM Operations"]
            col_map2 = get_col_map(ws2)
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                parent = clean_val(row[col_map2.get("parent", -1)]) if "parent" in col_map2 else ""
                if not parent: continue
                
                if parent not in boms:
                    # If an operation exists for a parent that wasn't in Materials, initialize it
                    boms[parent] = {"qty": 1.0, "items": [], "operations": []}
                    dependencies[parent] = set()
                
                op_name = clean_val(row[col_map2.get("operation", -1)]) if "operation" in col_map2 else ""
                if op_name:
                    sub = clean_val(row[col_map2.get("is_subcontracted", -1)]) if "is_subcontracted" in col_map2 else ""
                    boms[parent]["operations"].append({
                        "operation": op_name,
                        "workstation": clean_val(row[col_map2.get("workstation", -1)]),
                        "time_in_mins": flt(row[col_map2.get("op_time", -1)]),
                        "hour_rate": flt(row[col_map2.get("hour_rate", -1)]),
                        "is_subcontracted": 1 if sub.lower() in ["yes", "y", "1", "true"] else 0
                    })

        # Topological sort for bottom-up creation
        def topological_sort(dep_graph):
            visited = set()
            temp_mark = set()
            order = []
            
            def visit(node):
                if node in temp_mark: return # Circular dependency
                if node not in visited:
                    temp_mark.add(node)
                    for child in dep_graph.get(node, []):
                        if child in dep_graph: # Only visit if child is a parent BOM in this import
                            visit(child)
                    temp_mark.remove(node)
                    visited.add(node)
                    order.append(node)
                    
            for node in dep_graph:
                if node not in visited:
                    visit(node)
            return order

        creation_order = topological_sort(dependencies)
        
        summary = []
        for parent_code in creation_order:
            data = boms[parent_code]
            try:
                bom = frappe.new_doc("BOM")
                bom.item = parent_code
                bom.quantity = data["qty"]
                bom.is_active = 1
                bom.is_default = 1
                bom.with_operations = 1 if data["operations"] else 0
                
                for item_data in data["items"]:
                    bom.append("items", {
                        "item_code": item_data["item_code"],
                        "qty": item_data["qty"]
                    })
                
                for op_data in data["operations"]:
                    if not frappe.db.exists("Operation", op_data["operation"]):
                        new_op = frappe.new_doc("Operation")
                        new_op.operation = op_data["operation"]
                        new_op.insert(ignore_permissions=True)
                    
                    op_row = bom.append("operations", {
                        "operation": op_data["operation"],
                        "workstation": op_data["workstation"],
                        "time_in_mins": op_data["time_in_mins"],
                        "hour_rate": op_data["hour_rate"]
                    })
                    
                    if op_data["is_subcontracted"] and hasattr(op_row, "is_subcontracted"):
                        op_row.is_subcontracted = 1
                
                bom.insert(ignore_permissions=True)
                bom.submit()
                summary.append(f"✅ {parent_code}: BOM Created")
                
            except Exception as e:
                msg = str(e)
                if hasattr(e, 'message'): msg = e.message
                elif hasattr(e, 'args') and e.args: msg = e.args[0]
                from frappe.utils import strip_html
                msg = strip_html(str(msg))
                summary.append(f"❌ {parent_code}: {msg}")

        status = "Completed" if not any("❌" in log for log in summary) else "Failed"
        doc.db_set("status", status)
        doc.db_set("processing_log", "SUMMARY:\n" + "\n".join(summary))
        frappe.db.commit()

    except Exception as e:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Critical Error:\n{str(e)}")
        frappe.db.commit()
