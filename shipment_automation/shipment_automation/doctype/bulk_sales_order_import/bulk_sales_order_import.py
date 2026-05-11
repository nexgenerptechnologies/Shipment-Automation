import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date, validate_2_of_3


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Sales Order Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Sales Order Import"
    
    headers = [
        "Sales Order Number (Optional)", "Sales Order Date", "Customer Name", 
        "Customer PO Number", "Customer PO Date", "Item Code", "Item Name", 
        "Description", "Quantity", "Rate", "Delivery Date"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Sales_Order_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkSalesOrderImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_sales_order_import.bulk_sales_order_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("so_log", "⏳ Creating Sales Orders. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_sales_order_import.bulk_sales_order_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "so_num": ["Sales Order Number (Optional)", "Sales Order Number", "SO Number"],
        "so_date": ["Sales Order Date", "SO Date"],
        "customer": ["Customer Name", "Customer"],
        "cust_po": ["Customer PO Number"],
        "cust_po_date": ["Customer PO Date"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "delivery_date": ["Delivery Date"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Sales Order Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        today = nowdate()
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            so_id = str(row[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and row[col_map["so_num"]] else ""
            customer = str(row[col_map["customer"]]).strip() if col_map.get("customer") is not None else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None else ""
            
            so_date = parse_excel_date(row[col_map["so_date"]])
            delivery_date = parse_excel_date(row[col_map["delivery_date"]])

            if not customer or not frappe.db.exists("Customer", customer):
                row_errors.append(f"Customer '{customer}' not found.")
            
            if not item_code or not frappe.db.exists("Item", item_code):
                row_errors.append(f"Item '{item_code}' not found.")
            else:
                # 2-of-3 column match
                it_doc = frappe.get_doc("Item", item_code)
                if not validate_2_of_3(it_doc, item_code, item_name, description):
                    row_errors.append("2-of-3 match failed (Code/Name/Description).")

            if so_date and getdate(so_date) > getdate(today):
                row_errors.append(f"SO Date '{so_date}' is a future date.")

            if so_id and frappe.db.exists("Sales Order", so_id):
                row_errors.append(f"Duplicate SO Number '{so_id}' already exists.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Sales Order Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Group by SO ID if provided, else by Customer + SO Date
        so_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            so_id = str(row[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and row[col_map["so_num"]] else ""
            cust = str(row[col_map["customer"]]).strip()
            date = parse_excel_date(row[col_map["so_date"]]) or nowdate()
            
            group_key = so_id if so_id else (cust, date)
            if group_key not in so_groups: so_groups[group_key] = []
            so_groups[group_key].append(row)

        created = []
        for key, rows in so_groups.items():
            try:
                first = rows[0]
                customer = str(first[col_map["customer"]]).strip()
                so_id = str(first[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and first[col_map["so_num"]] else ""
                so_date = parse_excel_date(first[col_map["so_date"]]) or nowdate()
                
                so = frappe.new_doc("Sales Order")
                if so_id:
                    so.name = so_id
                
                so.customer = customer
                so.transaction_date = so_date
                so.po_no = str(first[col_map["cust_po"]]).strip() if col_map.get("cust_po") is not None else ""
                so.po_date = parse_excel_date(first[col_map["cust_po_date"]])
                
                for r in rows:
                    so.append("items", {
                        "item_code": str(r[col_map["item_code"]]).strip(),
                        "qty": flt(r[col_map["quantity"]]),
                        "rate": flt(r[col_map["rate"]]),
                        "delivery_date": parse_excel_date(r[col_map["delivery_date"]]) or so_date
                    })

                so.flags.ignore_permissions = True
                so.insert()
                # so.submit() # Optional: Typically SO is kept in Draft for review
                created.append(f"✅ {so.name}")
            except Exception as e:
                created.append(f"❌ Error creating SO: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("so_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("so_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
