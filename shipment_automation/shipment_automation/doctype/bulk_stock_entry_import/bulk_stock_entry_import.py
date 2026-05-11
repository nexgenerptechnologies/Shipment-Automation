import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date, validate_2_of_3


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Stock Entry Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Stock Entry Import"
    
    headers = [
        "Voucher ID (Link rows)", "Stock Entry Type", "Posting Date", 
        "Item Code", "Item Name", "Description", "Quantity", 
        "Source Warehouse", "Target Warehouse", "User Remark"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Stock_Entry_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkStockEntryImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_stock_entry_import.bulk_stock_entry_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("stock_log", "⏳ Creating Stock Entries. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_stock_entry_import.bulk_stock_entry_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "v_id": ["Voucher ID (Link rows)", "Voucher ID"],
        "type": ["Stock Entry Type"],
        "date": ["Posting Date"],
        "item": ["Item Code"],
        "item_name": ["Item Name"],
        "desc": ["Description"],
        "qty": ["Quantity", "Qty"],
        "src_wh": ["Source Warehouse"],
        "target_wh": ["Target Warehouse"],
        "remark": ["User Remark", "Remark"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Stock Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            s_type = str(row[col_map["type"]]).strip() if col_map.get("type") is not None else ""
            item_code = str(row[col_map["item"]]).strip() if col_map.get("item") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            desc = str(row[col_map["desc"]]).strip() if col_map.get("desc") is not None else ""
            src_wh = str(row[col_map["src_wh"]]).strip() if col_map.get("src_wh") is not None and row[col_map["src_wh"]] else ""
            target_wh = str(row[col_map["target_wh"]]).strip() if col_map.get("target_wh") is not None and row[col_map["target_wh"]] else ""
            
            posting_date = parse_excel_date(row[col_map["date"]])

            if not s_type: row_errors.append("Stock Entry Type is mandatory.")
            
            if not item_code or not frappe.db.exists("Item", item_code):
                row_errors.append(f"Item '{item_code}' not found.")
            else:
                it_doc = frappe.get_doc("Item", item_code)
                if not validate_2_of_3(it_doc, item_code, item_name, desc):
                    row_errors.append("2-of-3 match failed (Code/Name/Description).")

            if src_wh and not frappe.db.exists("Warehouse", src_wh):
                row_errors.append(f"Source Warehouse '{src_wh}' not found.")
            
            if target_wh and not frappe.db.exists("Warehouse", target_wh):
                row_errors.append(f"Target Warehouse '{target_wh}' not found.")

            if posting_date and getdate(posting_date) > getdate(nowdate()):
                row_errors.append(f"Posting Date '{posting_date}' is a future date.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Stock Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Group by Voucher ID
        groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            v_id = str(row[col_map["v_id"]]).strip() if col_map.get("v_id") is not None and row[col_map["v_id"]] else "SINGLE"
            if v_id not in groups: groups[v_id] = []
            groups[v_id].append(row)

        created = []
        for v_id, rows in groups.items():
            try:
                first = rows[0]
                se = frappe.new_doc("Stock Entry")
                se.stock_entry_type = str(first[col_map["type"]]).strip()
                se.posting_date = parse_excel_date(first[col_map["date"]])
                se.remarks = str(first[col_map["remark"]]).strip() if col_map.get("remark") is not None else f"Bulk Import {v_id}"
                
                # Global warehouses if applicable
                se.from_warehouse = str(first[col_map["src_wh"]]).strip() if col_map.get("src_wh") is not None and first[col_map["src_wh"]] else None
                se.to_warehouse = str(first[col_map["target_wh"]]).strip() if col_map.get("target_wh") is not None and first[col_map["target_wh"]] else None
                
                for r in rows:
                    se.append("items", {
                        "item_code": str(r[col_map["item"]]).strip(),
                        "qty": flt(r[col_map["qty"]]),
                        "s_warehouse": str(r[col_map["src_wh"]]).strip() if col_map.get("src_wh") is not None and r[col_map["src_wh"]] else se.from_warehouse,
                        "t_warehouse": str(r[col_map["target_wh"]]).strip() if col_map.get("target_wh") is not None and r[col_map["target_wh"]] else se.to_warehouse,
                        "basic_rate": 0 # ERPNext will auto-fetch moving average/FIFO rate on submission
                    })

                se.flags.ignore_permissions = True
                se.insert()
                # se.submit() # Draft for user review
                created.append(f"✅ {se.name}")
            except Exception as e:
                created.append(f"❌ Voucher '{v_id}': {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("stock_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("stock_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
