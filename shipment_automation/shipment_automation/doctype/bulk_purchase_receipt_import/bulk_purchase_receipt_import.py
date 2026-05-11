import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Purchase Receipt Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk PR Import Template"
    
    headers = [
        "Purchase Receipt Number", "Purchase Receipt Date", "Supplier Name", 
        "Purchase Order Number", "Item Code", "Item Name", 
        "Description", "Quantity", "Rate", "Line Number",
        "Purchase Invoice Number", "Purchase Invoice Date",
        "Bill of Entry Number", "Bill of Entry Date"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_PR_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkPurchaseReceiptImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_purchase_receipt_import.bulk_purchase_receipt_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("receipts_log", "⏳ Creating Purchase Receipts. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_purchase_receipt_import.bulk_purchase_receipt_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."

    @frappe.whitelist()
    def create_purchase_invoice_and_boe(self):
        """Processes the Excel file again to create Invoices and BOEs with custom IDs and Dates."""
        if self.status != "Completed":
             frappe.throw("Please process the Purchase Receipts first.")
        
        file_doc = frappe.get_doc("File", {"file_url": self.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Group by PR Number to ensure we only create one Invoice/BOE per PR
        pr_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            pr_id = str(row[col_map["pr_num"]]).strip()
            if pr_id not in pr_groups:
                pr_groups[pr_id] = row

        created = []
        for pr_id, row in pr_groups.items():
            try:
                # 1. Fetch the Purchase Receipt created in previous step
                if not frappe.db.exists("Purchase Receipt", pr_id):
                    created.append(f"❌ {pr_id}: Purchase Receipt not found.")
                    continue
                
                pr_doc = frappe.get_doc("Purchase Receipt", pr_id)
                
                # Check for mandatory custom IDs from Excel
                pi_id = str(row[col_map["pi_num"]]).strip() if col_map.get("pi_num") is not None and row[col_map["pi_num"]] else ""
                pi_date = parse_excel_date(row[col_map["pi_date"]]) if col_map.get("pi_date") is not None else None
                boe_id = str(row[col_map["boe_num"]]).strip() if col_map.get("boe_num") is not None and row[col_map["boe_num"]] else ""
                boe_date = parse_excel_date(row[col_map["boe_date"]]) if col_map.get("boe_date") is not None else None
                
                if not pi_id or not boe_id:
                    created.append(f"❌ {pr_id}: Purchase Invoice Number or BOE Number missing in Excel.")
                    continue

                # 2. Create Purchase Invoice
                if not frappe.db.exists("Purchase Invoice", pi_id):
                    from erpnext.buying.doctype.purchase_receipt.purchase_receipt import make_purchase_invoice
                    pi = make_purchase_invoice(pr_id)
                    pi.name = pi_id
                    pi.posting_date = pi_date or nowdate()
                    pi.bill_no = pi_id # Often used as Supplier Invoice No
                    
                    # Set Naming Series
                    pi_series = frappe.get_meta("Purchase Invoice").get_field("naming_series").options.split("\n")
                    pi.naming_series = pi_series[0]
                    for s in pi_series:
                        if pi_id.startswith(s.replace(".####", "").replace(".YY.", "").strip()):
                            pi.naming_series = s
                            break

                    pi.flags.ignore_permissions = True
                    pi.insert()
                    pi.submit()
                    invoice_name = pi.name
                else:
                    invoice_name = pi_id

                # 3. Create Bill of Entry
                if not frappe.db.exists("Bill of Entry", boe_id):
                    boe = frappe.new_doc("Bill of Entry")
                    boe.name = boe_id
                    boe.purchase_invoice = invoice_name
                    boe.posting_date = boe_date or nowdate()
                    boe.bill_of_entry_number = boe_id
                    boe.bill_of_entry_date = boe_date or nowdate()
                    boe.company = pr_doc.company
                    boe.supplier = pr_doc.supplier
                    boe.flags.ignore_permissions = True
                    boe.insert(ignore_mandatory=True)
                    # boe.submit() # Bill of Entry is usually a Custom DocType, depends if it has submit
                    boe_name = boe.name
                else:
                    boe_name = boe_id
                
                created.append(f"✅ {pr_id} -> Invoice {invoice_name}, BOE {boe_name}")
            except Exception as e:
                created.append(f"❌ {pr_id}: {str(e)}")
        
        frappe.db.commit()
        return {"summary": "\n".join(created)}


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "pr_num": ["Purchase Receipt Number", "PR Number"],
        "pr_date": ["Purchase Receipt Date", "PR Date"],
        "supplier": ["Supplier Name", "Supplier"],
        "po_num": ["Purchase Order Number", "PO Number"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "line_number": ["Line Number", "Line #"],
        "pi_num": ["Purchase Invoice Number", "PI Number", "Invoice No"],
        "pi_date": ["Purchase Invoice Date", "PI Date", "Invoice Date"],
        "boe_num": ["Bill of Entry Number", "BOE Number", "BOE No"],
        "boe_date": ["Bill of Entry Date", "BOE Date"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def find_po_item_by_line(po_num, item_code, line_val):
    """STRICTLY finds a PO Item by Line Number. Returns None if mismatch."""
    meta = frappe.get_meta("Purchase Order Item")
    available_fields = [f.fieldname for f in meta.fields]
    
    if "line_number" in available_fields:
        res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "line_number": line_val, "item_code": item_code}, "name")
        if res: return res
        
    if "custom_line_number" in available_fields:
        res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "custom_line_number": line_val, "item_code": item_code}, "name")
        if res: return res
        
    import re
    match = re.search(r'(\d+)$', po_num)
    if match:
        suffix = match.group(1)
        if line_val.startswith(f"{suffix}-"):
            try:
                idx_part = int(line_val.split("-")[-1])
                res = frappe.db.get_value("Purchase Order Item", {"parent": po_num, "idx": idx_part, "item_code": item_code}, "name")
                if res: return res
            except: pass
            
    return None


def parse_excel_date(date_val):
    """STRICTLY forces DD/MM/YYYY parsing even if Excel auto-converted it to MM/DD/YYYY."""
    if not date_val:
        return None
    
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        if date_val.day <= 12:
             try:
                 return datetime.date(date_val.year, date_val.day, date_val.month).strftime("%Y-%m-%d")
             except ValueError:
                 return date_val.strftime("%Y-%m-%d")
        return date_val.strftime("%Y-%m-%d")

    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    try:
        res = getdate(date_val)
        return res.strftime("%Y-%m-%d") if res else None
    except:
        return None


def run_validation(docname):
    doc = frappe.get_doc("Bulk Purchase Receipt Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        po_line_totals = {}
        po_line_rows = {}
        pr_number_map = {}
        today = nowdate()
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            
            pr_num = str(row[col_map["pr_num"]]).strip() if col_map.get("pr_num") is not None and row[col_map["pr_num"]] else ""
            raw_pr_date = row[col_map["pr_date"]] if col_map.get("pr_date") is not None else None
            supplier_name = str(row[col_map["supplier"]]).strip() if col_map.get("supplier") is not None else ""
            po_num = str(row[col_map["po_num"]]).strip() if col_map.get("po_num") is not None else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None else ""
            qty_exc = flt(row[col_map["quantity"]]) if col_map.get("quantity") is not None else 0
            rate_exc = flt(row[col_map["rate"]]) if col_map.get("rate") is not None else 0
            line_val = str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""

            pr_date_parsed = parse_excel_date(raw_pr_date)

            if not pr_num:
                row_errors.append("Purchase Receipt Number missing.")
            else:
                if pr_num not in pr_number_map:
                    pr_number_map[pr_num] = {"supplier": supplier_name, "po": po_num, "row": row_idx}
                else:
                    existing = pr_number_map[pr_num]
                    if existing["supplier"] != supplier_name or existing["po"] != po_num:
                        row_errors.append(f"Purchase Receipt Number '{pr_num}' is same for Purchase Order Number '{existing['po']}' & '{po_num}' and the suppliers '{existing['supplier']}' & '{supplier_name}' are different so Purchase Receipt can't created.")

                if frappe.db.exists("Purchase Receipt", pr_num):
                    row_errors.append(f"Duplicate PR Number '{pr_num}' already exists.")

            if pr_date_parsed and getdate(pr_date_parsed) > getdate(today):
                row_errors.append(f"Purchase Receipt Number '{pr_num}', Date is future Date, psl correct the Purchase Receipt Date.")

            if not po_num or not frappe.db.exists("Purchase Order", po_num):
                row_errors.append(f"PO '{po_num}' not found.")
            else:
                if pr_date_parsed:
                    po_date_obj = getdate(frappe.db.get_value("Purchase Order", po_num, "transaction_date"))
                    if getdate(pr_date_parsed) < po_date_obj:
                        row_errors.append(f"PR Date ({pr_date_parsed}) cannot be before PO Date ({po_date_obj.strftime('%Y-%m-%d')}).")

                po_supplier = frappe.db.get_value("Purchase Order", po_num, "supplier")
                if po_supplier != supplier_name:
                    row_errors.append(f"Supplier mismatch: PO is for '{po_supplier}', Excel has '{supplier_name}'.")

                po_item_name = find_po_item_by_line(po_num, item_code, line_val) if line_val else frappe.db.get_value("Purchase Order Item", {"parent": po_num, "item_code": item_code}, "name")
                
                if not po_item_name:
                    row_errors.append(f"Item/Line '{line_val or item_code}' not found in PO '{po_num}'.")
                else:
                    pi = frappe.get_doc("Purchase Order Item", po_item_name)
                    total_key = (po_num, item_code, line_val)
                    if total_key not in po_line_totals:
                        po_line_totals[total_key] = 0.0
                        po_line_rows[total_key] = []
                    po_line_totals[total_key] += qty_exc
                    po_line_rows[total_key].append(str(row_idx))
                    
                    if po_line_totals[total_key] > flt(pi.qty) + 0.0000001:
                        rows_str = " & ".join(po_line_rows[total_key])
                        row_errors.append(f"total sum {po_line_totals[total_key]} of row number {rows_str} are more than Purchase Order Line Quantity {line_val}")
                    
                    if abs(flt(pi.rate) - flt(rate_exc)) > 0.0000001:
                        row_errors.append(f"Rate mismatch: Excel {rate_exc} vs PO {pi.rate}")

                    score = 0
                    if pi.item_code == item_code: score += 1
                    if pi.item_name == item_name: score += 1
                    if pi.description == description: score += 1
                    if score < 2:
                        row_errors.append("2-of-3 column match failed (Code/Name/Description).")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", f"❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Purchase Receipt Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        pr_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            pr_id = str(row[col_map["pr_num"]]).strip()
            pr_date = parse_excel_date(row[col_map["pr_date"]]) or "no-date"
            supplier = str(row[col_map["supplier"]]).strip()
            
            group_key = (pr_id, pr_date, supplier)
            if group_key not in pr_groups: pr_groups[group_key] = []
            pr_groups[group_key].append(row)

        created_receipts = []
        for (pr_id, pr_date_str, supplier), rows in pr_groups.items():
            first_row = rows[0]
            po_num_first = str(first_row[col_map["po_num"]]).strip()
            pr_date_parsed = parse_excel_date(first_row[col_map["pr_date"]])
            
            po_header = frappe.db.get_value("Purchase Order", po_num_first, ["company", "currency", "conversion_rate"], as_dict=True)
            if frappe.db.exists("Purchase Receipt", pr_id): continue

            pr = frappe.new_doc("Purchase Receipt")
            pr.name = pr_id
            
            available_series = frappe.get_meta("Purchase Receipt").get_field("naming_series").options.split("\n")
            matched_series = available_series[0]
            for series in available_series:
                prefix = series.replace(".####", "").replace(".YY.", "").replace(".YYYY.", "").strip()
                if prefix and pr_id.startswith(prefix):
                    matched_series = series
                    break
            
            pr.update({
                "naming_series": matched_series,
                "supplier": supplier,
                "company": po_header.company or frappe.db.get_single_value("Global Defaults", "default_company"),
                "currency": po_header.currency,
                "conversion_rate": po_header.conversion_rate or 1.0,
                "set_posting_time": 1,
                "posting_date": pr_date_parsed if pr_date_parsed else nowdate(),
                "posting_time": "00:00:00"
            })

            for row in rows:
                p_num = str(row[col_map["po_num"]]).strip()
                i_code = str(row[col_map["item_code"]]).strip()
                l_val = str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""
                
                po_item_n = find_po_item_by_line(p_num, i_code, l_val) if l_val else frappe.db.get_value("Purchase Order Item", {"parent": p_num, "item_code": item_code}, "name")
                if not po_item_n: continue
                
                po_item = frappe.get_doc("Purchase Order Item", po_item_n)
                pr_item = pr.append("items", {
                    "item_code": i_code,
                    "qty": flt(row[col_map["quantity"]]),
                    "rate": flt(row[col_map["rate"]]),
                    "purchase_order": p_num,
                    "purchase_order_item": po_item.name,
                    "warehouse": po_item.warehouse
                })
                pr_item.run_method("set_missing_values")
                l_field = "line_number" if hasattr(pr_item, "line_number") else "custom_line_number"
                if l_val: setattr(pr_item, l_field, l_val)

            pr.run_method("set_missing_values")
            pr.run_method("calculate_taxes_and_totals")
            pr.flags.ignore_permissions = True
            
            pr.db_insert()
            for child in pr.get_all_children(): child.db_insert()
            pr.run_method("on_update")
            created_receipts.append(pr.name)

        doc.db_set("status", "Completed")
        doc.db_set("receipts_log", "SUMMARY:\n" + "\n".join([f"✅ {name}" for name in created_receipts]))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("receipts_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
