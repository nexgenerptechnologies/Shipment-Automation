import frappe
from frappe.model.document import Document
from frappe.utils import flt, nowdate, getdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date, validate_2_of_3


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Delivery Note Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Delivery Note Import"
    
    headers = [
        "Delivery Note Number (Optional)", "Posting Date", "Customer Name", 
        "Sales Order Number", "Item Code", "Item Name", 
        "Description", "Quantity", "Rate", "Warehouse"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Delivery_Note_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkDeliveryNoteImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_delivery_note_import.bulk_delivery_note_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("dn_log", "⏳ Creating Delivery Notes. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_delivery_note_import.bulk_delivery_note_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "dn_num": ["Delivery Note Number (Optional)", "Delivery Note Number", "DN Number"],
        "posting_date": ["Posting Date"],
        "customer": ["Customer Name", "Customer"],
        "so_num": ["Sales Order Number", "SO Number"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "warehouse": ["Warehouse"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Delivery Note Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        today = nowdate()
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            dn_id = str(row[col_map["dn_num"]]).strip() if col_map.get("dn_num") is not None and row[col_map["dn_num"]] else ""
            customer = str(row[col_map["customer"]]).strip() if col_map.get("customer") is not None else ""
            so_num = str(row[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and row[col_map["so_num"]] else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map.get("item_code") is not None else ""
            item_name = str(row[col_map["item_name"]]).strip() if col_map.get("item_name") is not None else ""
            description = str(row[col_map["description"]]).strip() if col_map.get("description") is not None else ""
            warehouse = str(row[col_map["warehouse"]]).strip() if col_map.get("warehouse") is not None and row[col_map["warehouse"]] else ""
            
            posting_date = parse_excel_date(row[col_map["posting_date"]])

            if not customer or not frappe.db.exists("Customer", customer):
                row_errors.append(f"Customer '{customer}' not found.")
            
            if so_num and not frappe.db.exists("Sales Order", so_num):
                row_errors.append(f"Sales Order '{so_num}' not found.")
            
            if not item_code or not frappe.db.exists("Item", item_code):
                row_errors.append(f"Item '{item_code}' not found.")
            else:
                # 2-of-3 column match
                it_doc = frappe.get_doc("Item", item_code)
                if not validate_2_of_3(it_doc, item_code, item_name, description):
                    row_errors.append("2-of-3 match failed (Code/Name/Description).")

            if warehouse and not frappe.db.exists("Warehouse", warehouse):
                row_errors.append(f"Warehouse '{warehouse}' not found.")

            if posting_date and getdate(posting_date) > getdate(today):
                row_errors.append(f"Posting Date '{posting_date}' is a future date.")

            if dn_id and frappe.db.exists("Delivery Note", dn_id):
                row_errors.append(f"Duplicate Delivery Note Number '{dn_id}' already exists.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Delivery Note Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        # Group by DN ID if provided, else by Customer + Posting Date
        dn_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            dn_id = str(row[col_map["dn_num"]]).strip() if col_map.get("dn_num") is not None and row[col_map["dn_num"]] else ""
            cust = str(row[col_map["customer"]]).strip()
            date = parse_excel_date(row[col_map["posting_date"]]) or nowdate()
            
            group_key = dn_id if dn_id else (cust, date)
            if group_key not in dn_groups: dn_groups[group_key] = []
            dn_groups[group_key].append(row)

        created = []
        for key, rows in dn_groups.items():
            try:
                first = rows[0]
                customer = str(first[col_map["customer"]]).strip()
                dn_id = str(first[col_map["dn_num"]]).strip() if col_map.get("dn_num") is not None and first[col_map["dn_num"]] else ""
                posting_date = parse_excel_date(first[col_map["posting_date"]]) or nowdate()
                
                dn = frappe.new_doc("Delivery Note")
                if dn_id:
                    dn.name = dn_id
                
                dn.customer = customer
                dn.posting_date = posting_date
                
                for r in rows:
                    so_num = str(r[col_map["so_num"]]).strip() if col_map.get("so_num") is not None and r[col_map["so_num"]] else ""
                    item_code = str(r[col_map["item_code"]]).strip()
                    qty = flt(r[col_map["quantity"]])
                    rate = flt(r[col_map["rate"]])
                    wh = str(r[col_map["warehouse"]]).strip() if col_map.get("warehouse") is not None and r[col_map["warehouse"]] else ""
                    
                    item_row = dn.append("items", {
                        "item_code": item_code,
                        "qty": qty,
                        "rate": rate,
                        "against_sales_order": so_num,
                        "warehouse": wh
                    })
                    
                    if so_num:
                        # Pull specific details from Sales Order if it exists
                        so_item = frappe.db.get_value("Sales Order Item", 
                            {"parent": so_num, "item_code": item_code}, 
                            ["name", "rate"], as_dict=True)
                        if so_item:
                            item_row.so_detail = so_item.name
                            if not rate: # If rate not in excel, pull from SO
                                item_row.rate = so_item.rate

                dn.flags.ignore_permissions = True
                dn.insert()
                # dn.submit() # Optional
                created.append(f"✅ {dn.name}")
            except Exception as e:
                created.append(f"❌ Error creating DN: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("dn_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("dn_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
