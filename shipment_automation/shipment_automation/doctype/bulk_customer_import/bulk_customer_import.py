import frappe
from frappe.model.document import Document
from frappe.utils import nowdate
import openpyxl
from io import BytesIO
import datetime
from shipment_automation.shipment_automation.utils import parse_excel_date


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Customer Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Customer Import"
    
    headers = [
        "Customer ID (Leave blank for Auto)", "Naming Series", "Customer Name", 
        "Customer Group", "Territory", "GST Category", "GSTIN",
        "Address Line 1", "City", "State", "Pincode", "Country",
        "Contact Person Name", "Email Address", "Mobile Number"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Customer_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkCustomerImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_customer_import.bulk_customer_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_processing(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Customers, Addresses & Contacts. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_customer_import.bulk_customer_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "id": ["Customer ID (Leave blank for Auto)", "Customer ID"],
        "series": ["Naming Series"],
        "name": ["Customer Name"],
        "group": ["Customer Group"],
        "territory": ["Territory"],
        "gst_cat": ["GST Category"],
        "gstin": ["GSTIN"],
        "street": ["Address Line 1", "Street Address"],
        "city": ["City"],
        "state": ["State"],
        "pincode": ["Pincode"],
        "country": ["Country"],
        "contact_name": ["Contact Person Name"],
        "email": ["Email Address"],
        "mobile": ["Mobile Number"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Customer Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            manual_id = str(row[col_map["id"]]).strip() if col_map.get("id") is not None and row[col_map["id"]] else ""
            customer_name = str(row[col_map["name"]]).strip() if col_map.get("name") is not None else ""
            gstin = str(row[col_map["gstin"]]).strip() if col_map.get("gstin") is not None and row[col_map["gstin"]] else ""
            gst_cat = str(row[col_map["gst_cat"]]).strip() if col_map.get("gst_cat") is not None else ""

            if not customer_name:
                row_errors.append("Customer Name is mandatory.")

            # Check Duplicates
            if manual_id and frappe.db.exists("Customer", manual_id):
                row_errors.append(f"Manual ID '{manual_id}' already exists.")
            
            if not manual_id and frappe.db.exists("Customer", customer_name):
                row_errors.append(f"Customer with name '{customer_name}' already exists.")

            if gstin and frappe.db.exists("Customer", {"gstin": gstin}):
                 row_errors.append(f"GSTIN '{gstin}' is already registered to another customer.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Customer Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        created = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            manual_id = str(row[col_map["id"]]).strip() if col_map.get("id") is not None and row[col_map["id"]] else ""
            series = str(row[col_map["series"]]).strip() if col_map.get("series") is not None and row[col_map["series"]] else ""
            group = str(row[col_map["group"]]).strip() if col_map.get("group") is not None and row[col_map["group"]] else ""
            
            # Check if group is a 'Group' type
            if group and frappe.db.get_value("Customer Group", group, "is_group"):
                created.append(f"❌ {customer_name}: Cannot select a Group type Customer Group ({group}). Please select a non-group Customer Group.")
                continue
            
            territory = str(row[col_map["territory"]]).strip() if col_map.get("territory") is not None and row[col_map["territory"]] else ""
            
            # Check if territory is a 'Group' type
            if territory and frappe.db.get_value("Territory", territory, "is_group"):
                created.append(f"❌ {customer_name}: Cannot select a Group type Territory ({territory}). Please select a non-group Territory.")
                continue
            gst_cat = str(row[col_map["gst_cat"]]).strip() if col_map.get("gst_cat") is not None else ""
            gstin = str(row[col_map["gstin"]]).strip() if col_map.get("gstin") is not None and row[col_map["gstin"]] else ""
            
            try:
                # Use a transaction block (implicitly via insert/db_insert)
                # But we catch errors at each step and throw to prevent partial creation
                
                # 1. Create Customer
                c_doc = frappe.new_doc("Customer")
                
                # Naming Logic: Manual ID > Naming Series > Default
                if manual_id:
                    c_doc.name = manual_id
                    c_doc.flags.ignore_mandatory = True # To allow manual name setting
                elif series:
                    c_doc.naming_series = series
                
                c_doc.customer_name = customer_name
                c_doc.customer_group = group
                c_doc.territory = territory
                c_doc.gst_category = gst_cat
                c_doc.gstin = gstin
                c_doc.insert(ignore_permissions=True)
                
                # Verification Check: Ensure name matches if manual_id was provided
                if manual_id and c_doc.name != manual_id:
                     # Force update name via database if Buying Settings override it
                     frappe.db.set_value("Customer", c_doc.name, "name", manual_id, update_modified=False)
                     c_doc.name = manual_id
                
                # 2. Create Address (Mandatory check: Line 1, State, Country)
                street = str(row[col_map["street"]]).strip() if col_map.get("street") is not None and row[col_map["street"]] else ""
                state = str(row[col_map["state"]]).strip() if col_map.get("state") is not None and row[col_map["state"]] else ""
                country = str(row[col_map["country"]]).strip() if col_map.get("country") is not None and row[col_map["country"]] else "India"

                if street or state: # If any address info is provided, enforce mandatory fields
                    if not street or not state or not country:
                        frappe.throw(f"Address creation failed: Address Line 1, State and Country are mandatory.")
                        
                    addr = frappe.new_doc("Address")
                    addr.address_title = customer_name
                    addr.address_type = "Billing"
                    addr.address_line1 = street
                    addr.city = str(row[col_map["city"]]).strip() if col_map.get("city") is not None else ""
                    addr.state = state
                    
                    # Validate Pincode (6 digits, not starting with 0 for India)
                    pincode = str(row[col_map["pincode"]]).strip() if col_map.get("pincode") is not None else ""
                    if pincode and (len(pincode) != 6 or pincode.startswith("0")):
                         pincode = "" # Clear invalid pincode to prevent error
                    
                    addr.pincode = pincode
                    addr.country = country
                    addr.append("links", {"link_doctype": "Customer", "link_name": c_doc.name})
                    addr.insert(ignore_permissions=True)

                # 3. Create Contact
                c_person = str(row[col_map["contact_name"]]).strip() if col_map.get("contact_name") is not None and row[col_map["contact_name"]] else ""
                email = str(row[col_map["email"]]).strip() if col_map.get("email") is not None and row[col_map["email"]] else ""
                mobile = str(row[col_map["mobile"]]).strip() if col_map.get("mobile") is not None and row[col_map["mobile"]] else ""

                if c_person:
                    con = frappe.new_doc("Contact")
                    con.first_name = c_person
                    # Link to the newly created customer name
                    con.append("links", {"link_doctype": "Customer", "link_name": c_doc.name})
                    con.flags.ignore_permissions = True
                    con.insert()
                    
                    if email:
                         con.append("email_ids", {"email_id": email, "is_primary": 1})
                    
                    if mobile:
                         con.append("phone_nos", {"phone_number": mobile, "is_primary": 1})
                    
                    con.save(ignore_permissions=True)
                
                # 4. Final Verification Message
                msg = f"Customer {customer_name} created successfully"
                details = []
                if street or state: details.append("Address")
                if c_person: details.append("Contact")
                if email: details.append(f"Email({email})")
                if mobile: details.append(f"Mobile({mobile})")
                
                if details:
                    msg += f" with " + ", ".join(details)
                
                created.append(f"✅ {msg}")
            except Exception as e:
                # If anything fails, we rollback the specific customer creation
                frappe.db.rollback()
                created.append(f"❌ {customer_name}: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("processing_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
