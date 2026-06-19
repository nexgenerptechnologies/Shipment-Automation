import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, nowdate
import openpyxl
from io import BytesIO
import datetime


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk JE Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk JE Import Template"
    
    headers = [
        "Voucher ID (Link rows)", "Entry Type", "Posting Date", "Account", 
        "Party", "Party Type", "Debit", "Credit", 
        "Bill No", "Bill Date", "Due Date", "User Remark"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Journal_Entry_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkJournalEntryImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_journal_entry_import.bulk_journal_entry_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("je_log", "⏳ Creating Journal Entries. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_journal_entry_import.bulk_journal_entry_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "v_id": ["Voucher ID (Link rows)", "Voucher ID"],
        "v_type": ["Entry Type", "Voucher Type"],
        "posting_date": ["Posting Date"],
        "account": ["Account (Accounting Entries)", "Account"],
        "party": ["Party (Accounting Entries)", "Party", "Party (Optional)"],
        "party_type": ["Party Type (Accounting Entries)", "Party Type", "Party Type (Optional)"],
        "debit": ["Debit (Accounting Entries)", "Debit", "Debit (INR)", "Debit (USD)"],
        "credit": ["Credit (Accounting Entries)", "Credit", "Credit (INR)", "Credit (USD)"],
        "bill_no": ["Bill No"],
        "bill_date": ["Bill Date"],
        "due_date": ["Due Date"],
        "remark": ["User Remark", "Remark"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def parse_excel_date(date_val):
    if not date_val:
        return None
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime("%Y-%m-%d")
    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"]:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    try:
        res = getdate(date_val)
        return res.strftime("%Y-%m-%d") if res else None
    except:
        return None


def resolve_party_id(party_type, party_name):
    if not party_type or not party_name:
        return party_name
    party_field = frappe.scrub(party_type) + "_name"
    party_id = frappe.db.get_value(party_type, {party_field: party_name}, "name")
    if not party_id:
        party_id = frappe.db.get_value(party_type, {"name": party_name}, "name")
    return party_id or party_name

def get_temporary_opening_account():
    # Try to find an account with 'Temporary Opening' in its name
    acc = frappe.db.get_value("Account", {"account_name": ["like", "%Temporary Opening%"]}, "name")
    if acc: return acc
    return "Temporary Opening"

def run_validation(docname):
    doc = frappe.get_doc("Bulk Journal Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        info_logs = []
        
        # Check for missing mandatory columns
        missing_cols = []
        if "debit" not in col_map: missing_cols.append("Debit")
        if "credit" not in col_map: missing_cols.append("Credit")
        if "account" not in col_map: missing_cols.append("Account")
        if "v_type" not in col_map: missing_cols.append("Entry Type")
        
        if missing_cols:
            errors.append(f"❌ Missing required columns in Excel header: {', '.join(missing_cols)}")
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\\n\\n" + "\\n".join(errors))
            frappe.db.commit()
            return
            
        voucher_groups = {}
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            v_id_val = str(row[col_map["v_id"]]).strip() if col_map.get("v_id") is not None and row[col_map["v_id"]] else ""
            v_type = str(row[col_map["v_type"]]).strip() if col_map.get("v_type") is not None else ""
            account = str(row[col_map["account"]]).strip() if col_map.get("account") is not None and row[col_map["account"]] else ""
            p_type = str(row[col_map["party_type"]]).strip() if col_map.get("party_type") is not None and row[col_map["party_type"]] else ""
            party_val = str(row[col_map["party"]]).strip() if col_map.get("party") is not None and row[col_map["party"]] else ""
            
            party = resolve_party_id(p_type, party_val) if p_type and party_val else ""
            
            debit = flt(row[col_map["debit"]]) if col_map.get("debit") is not None else 0.0
            credit = flt(row[col_map["credit"]]) if col_map.get("credit") is not None else 0.0
            
            bill_no = str(row[col_map["bill_no"]]).strip() if col_map.get("bill_no") is not None and row[col_map["bill_no"]] else ""
            bill_date = row[col_map["bill_date"]] if col_map.get("bill_date") is not None else None
            due_date = row[col_map["due_date"]] if col_map.get("due_date") is not None else None
            remark = str(row[col_map["remark"]]).strip() if col_map.get("remark") is not None and row[col_map["remark"]] else ""
            
            posting_date = parse_excel_date(row[col_map["posting_date"]]) if col_map.get("posting_date") is not None else ""

            # Flexible Grouping: Use Voucher ID if present, else fallback to auto-grouping
            if v_id_val:
                v_id = v_id_val
                display = f"Voucher '{v_id}'"
            else:
                v_id = f"{v_type}_{posting_date}_{remark}"
                display = f"Group '{v_type} on {posting_date}'"

            if not v_type: row_errors.append("Entry Type is mandatory.")
            
            if not account:
                pass # If user leaves account blank, we might just assume it's entirely Party based, or error out
            elif not frappe.db.exists("Account", account):
                row_errors.append(f"Account '{account}' not found.")
            else:
                acc_type = frappe.db.get_value("Account", account, "account_type")
                is_party_acc = acc_type in ["Payable", "Receivable"]
                if bill_no and not is_party_acc:
                    row_errors.append(f"Bill No '{bill_no}' cannot be used on standard account '{account}'. It requires a Payable/Receivable account.")

            if p_type and party_val:
                if not frappe.db.exists(p_type, party):
                    row_errors.append(f"Party '{party_val}' of type '{p_type}' not found.")
            
            if not posting_date:
                row_errors.append("Invalid Posting Date (Use DD/MM/YYYY).")
            elif getdate(posting_date) > getdate(nowdate()):
                row_errors.append(f"Posting Date '{posting_date}' is a future date.")
            
            if bill_date:
                parsed_bd = parse_excel_date(bill_date)
                if not parsed_bd: row_errors.append("Invalid Bill Date (Use DD/MM/YYYY).")
            
            if due_date:
                parsed_dd = parse_excel_date(due_date)
                if not parsed_dd: row_errors.append("Invalid Due Date (Use DD/MM/YYYY).")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            
            # Sum debits/credits per voucher to check balance later
            if v_id not in voucher_groups: voucher_groups[v_id] = {"d": 0, "c": 0, "display": display}
            voucher_groups[v_id]["d"] += debit
            voucher_groups[v_id]["c"] += credit

        # Final Balance Check (Auto-Balancing Logic)
        for vid, sums in voucher_groups.items():
            diff = sums["d"] - sums["c"]
            if abs(diff) > 0.01:
                info_logs.append(f"ℹ️ {sums['display']} is out of balance. Auto-Balancing: Difference of {abs(diff):.2f} will be posted to Temporary Opening.")

        if not errors:
            doc.db_set("status", "Validated")
            log_msg = f"✅ All entries validated successfully."
            if info_logs:
                log_msg += "\\n\\n" + "\\n".join(info_logs)
            doc.db_set("validation_log", log_msg)
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\\n\\n" + "\\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Journal Entry Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        if "debit" not in col_map or "credit" not in col_map:
            raise Exception("Missing Debit or Credit column. Please fix the Excel header.")
            
        # Grouping
        groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            v_id_val = str(row[col_map["v_id"]]).strip() if col_map.get("v_id") is not None and row[col_map["v_id"]] else ""
            v_type = str(row[col_map["v_type"]]).strip() if col_map.get("v_type") is not None else ""
            remark = str(row[col_map["remark"]]).strip() if col_map.get("remark") is not None and row[col_map["remark"]] else ""
            posting_date = parse_excel_date(row[col_map["posting_date"]]) if col_map.get("posting_date") is not None else ""
            
            if v_id_val:
                v_id = v_id_val
            else:
                v_id = f"{v_type}_{posting_date}_{remark}"
            
            if v_id not in groups: groups[v_id] = []
            groups[v_id].append(row)

        created = []
        for v_id, rows in groups.items():
            try:
                first = rows[0]
                je = frappe.new_doc("Journal Entry")
                je.voucher_type = str(first[col_map["v_type"]]).strip()
                je.posting_date = parse_excel_date(first[col_map["posting_date"]])
                remark_val = str(first[col_map["remark"]]).strip() if col_map.get("remark") is not None and str(first[col_map["remark"]]).strip() else f"Bulk Import {je.voucher_type}"
                je.user_remark = remark_val
                
                total_debit = 0.0
                total_credit = 0.0
                
                for r in rows:
                    p_type = str(r[col_map["party_type"]]).strip() if col_map.get("party_type") is not None and r[col_map["party_type"]] else None
                    party_val = str(r[col_map["party"]]).strip() if col_map.get("party") is not None and r[col_map["party"]] else None
                    party = resolve_party_id(p_type, party_val) if p_type and party_val else None
                    
                    debit_val = flt(r[col_map["debit"]]) if col_map.get("debit") is not None else 0.0
                    credit_val = flt(r[col_map["credit"]]) if col_map.get("credit") is not None else 0.0
                    
                    total_debit += debit_val
                    total_credit += credit_val
                    
                    acc_row = {
                        "account": str(r[col_map["account"]]).strip() if r[col_map["account"]] else "",
                        "party_type": p_type,
                        "party": party,
                        "debit_in_account_currency": debit_val,
                        "credit_in_account_currency": credit_val
                    }
                    
                    bill_no = str(r[col_map["bill_no"]]).strip() if col_map.get("bill_no") is not None and r[col_map["bill_no"]] else ""
                    if bill_no:
                        if je.voucher_type == "Opening Entry":
                            # For Opening Entries, the original invoice doesn't exist in ERPNext yet.
                            # So we cannot link it via reference_type/name (it would throw a Link Error).
                            acc_row["user_remark"] = f"Bill No: {bill_no}"
                        else:
                            acc_row["reference_type"] = "Purchase Invoice" if credit_val > 0 else "Sales Invoice" 
                            acc_row["reference_name"] = bill_no
                        
                        acc_row["bill_no"] = bill_no
                        
                    bill_date = r[col_map["bill_date"]] if col_map.get("bill_date") is not None else None
                    if bill_date:
                        bd = parse_excel_date(bill_date)
                        if bd:
                            acc_row["reference_date"] = bd
                            acc_row["bill_date"] = bd
                            
                    due_date = r[col_map["due_date"]] if col_map.get("due_date") is not None else None
                    if due_date:
                        dd = parse_excel_date(due_date)
                        if dd:
                            acc_row["reference_due_date"] = dd
                            acc_row["due_date"] = dd

                    je.append("accounts", acc_row)

                # Auto Balancing Logic
                diff = total_debit - total_credit
                if abs(diff) > 0.01:
                    balancing_account = get_temporary_opening_account()
                    acc_row = {
                        "account": balancing_account,
                        "debit_in_account_currency": abs(diff) if diff < 0 else 0.0,
                        "credit_in_account_currency": diff if diff > 0 else 0.0
                    }
                    je.append("accounts", acc_row)

                je.flags.ignore_permissions = True
                je.insert()
                created.append(f"✅ {je.name}")
            except Exception as e:
                msg = str(e)
                if hasattr(e, 'message'): msg = e.message
                elif hasattr(e, 'args') and e.args: msg = e.args[0]
                from frappe.utils import strip_html
                msg = strip_html(str(msg))
                created.append(f"❌ {je.voucher_type} on {je.posting_date}: {msg}")

        status = "Completed" if not any("❌" in log for log in created) else "Failed"
        doc.db_set("status", status)
        doc.db_set("je_log", "SUMMARY:\\n" + "\\n".join(created))
        frappe.db.commit()
    except Exception as e:
        doc.db_set("status", "Failed")
        doc.db_set("je_log", f"❌ Critical Error:\\n{str(e)}")
        frappe.db.commit()
