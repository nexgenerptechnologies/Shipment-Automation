import frappe
from frappe.model.document import Document
from frappe.utils import nowdate, getdate
import openpyxl
from io import BytesIO
import datetime
import zipfile
import os
from shipment_automation.shipment_automation.utils import parse_excel_date


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk Employee Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Employee Import"
    
    headers = [
        "Employee ID", "Employee Name", "Gender", "Date of Birth", 
        "Date of Joining", "Company", "Department", "Designation", 
        "Branch", "Reports to (Emp ID)", "Personal Email", "Mobile Number"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_Employee_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkEmployeeImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_employee_import.bulk_employee_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("emp_log", "⏳ Creating Employees & Processing Images. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_employee_import.bulk_employee_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "id": ["Employee ID"],
        "name": ["Employee Name"],
        "gender": ["Gender"],
        "dob": ["Date of Birth"],
        "doj": ["Date of Joining"],
        "company": ["Company"],
        "dept": ["Department"],
        "desig": ["Designation"],
        "branch": ["Branch"],
        "reports_to": ["Reports to (Emp ID)"],
        "email": ["Personal Email"],
        "mobile": ["Mobile Number"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def run_validation(docname):
    doc = frappe.get_doc("Bulk Employee Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            row_errors = []
            emp_id = str(row[col_map["id"]]).strip() if col_map.get("id") is not None else ""
            emp_name = str(row[col_map["name"]]).strip() if col_map.get("name") is not None else ""
            company = str(row[col_map["company"]]).strip() if col_map.get("company") is not None else ""
            
            dob = parse_excel_date(row[col_map["dob"]])
            doj = parse_excel_date(row[col_map["doj"]])

            if not emp_id: row_errors.append("Employee ID is mandatory.")
            elif frappe.db.exists("Employee", emp_id): row_errors.append(f"Employee ID '{emp_id}' already exists.")
            
            if not emp_name: row_errors.append("Employee Name is mandatory.")
            
            if not company or not frappe.db.exists("Company", company):
                row_errors.append(f"Company '{company}' not found.")

            if doj and getdate(doj) > getdate(nowdate()):
                row_errors.append(f"DOJ '{doj}' cannot be a future date.")

            if row_errors:
                errors.append(f"Row {row_idx} ❌ " + " | ".join(row_errors))
            else:
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} employee(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk Employee Import", docname)
    temp_dir = None
    try:
        # 1. Unzip Images if provided
        img_map = {}
        if doc.image_zip:
            zip_file_doc = frappe.get_doc("File", {"file_url": doc.image_zip})
            temp_dir = frappe.get_site_path("private", "files", f"temp_zip_{doc.name}")
            if not os.path.exists(temp_dir): os.makedirs(temp_dir)
            
            with zipfile.ZipFile(zip_file_doc.get_full_path(), 'r') as zip_ref:
                zip_ref.extractall(temp_dir)
            
            for f in os.listdir(temp_dir):
                name, ext = os.path.splitext(f)
                if ext.lower() in ['.jpg', '.jpeg', '.png']:
                    img_map[name.strip()] = os.path.join(temp_dir, f)

        # 2. Process Excel
        file_doc = frappe.get_doc("File", {"file_url": doc.excel_file})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        created = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            
            emp_id = str(row[col_map["id"]]).strip()
            try:
                emp = frappe.new_doc("Employee")
                emp.employee = emp_id
                emp.first_name = str(row[col_map["name"]]).strip()
                emp.gender = str(row[col_map["gender"]]).strip() if col_map.get("gender") is not None else ""
                emp.date_of_birth = parse_excel_date(row[col_map["dob"]])
                emp.date_of_joining = parse_excel_date(row[col_map["doj"]])
                emp.company = str(row[col_map["company"]]).strip()
                emp.department = str(row[col_map["dept"]]).strip() if col_map.get("dept") is not None else ""
                emp.designation = str(row[col_map["desig"]]).strip() if col_map.get("desig") is not None else ""
                emp.branch = str(row[col_map["branch"]]).strip() if col_map.get("branch") is not None else ""
                emp.reports_to = str(row[col_map["reports_to"]]).strip() if col_map.get("reports_to") is not None else ""
                emp.personal_email = str(row[col_map["email"]]).strip() if col_map.get("email") is not None else ""
                emp.cell_number = str(row[col_map["mobile"]]).strip() if col_map.get("mobile") is not None else ""
                
                # 3. Handle Image
                if emp_id in img_map:
                    img_path = img_map[emp_id]
                    with open(img_path, "rb") as f:
                        content = f.read()
                    
                    filename = os.path.basename(img_path)
                    f_doc = frappe.new_doc("File")
                    f_doc.file_name = filename
                    f_doc.content = content
                    f_doc.attached_to_doctype = "Employee"
                    f_doc.attached_to_name = emp_id
                    f_doc.is_private = 0
                    f_doc.save(ignore_permissions=True)
                    
                    emp.image = f_doc.file_url

                emp.flags.ignore_permissions = True
                emp.insert()
                created.append(f"✅ {emp_id}")
            except Exception as e:
                created.append(f"❌ {emp_id}: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("emp_log", "SUMMARY:\n" + "\n".join(created))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("emp_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
    finally:
        # Cleanup temp zip folder
        if temp_dir and os.path.exists(temp_dir):
            import shutil
            shutil.rmtree(temp_dir)
