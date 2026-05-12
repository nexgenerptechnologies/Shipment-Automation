import frappe
import openpyxl
import zipfile
import os
from frappe.model.document import Document
from frappe.utils import flt, get_site_path
from frappe.utils.file_manager import save_file

class BulkItemImport(Document):
    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validating Excel data...")
        frappe.db.commit()
        frappe.enqueue("shipment_automation.shipment_automation.doctype.bulk_item_import.bulk_item_import.run_validation", docname=self.name)

    @frappe.whitelist()
    def start_processing(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Items and linking images. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue("shipment_automation.shipment_automation.doctype.bulk_item_import.bulk_item_import.run_processing", docname=self.name)

    @frappe.whitelist()
    def download_template(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Item Import Template"
        
        headers = [
            "Item Code", "Item Name", "Item Group", "Default Unit of Measure", 
            "Opening Stock", "Valuation Rate", "Standard Selling Rate", "Is Stock Item (Yes/No)"
        ]
        ws.append(headers)
        
        # Add sample data
        ws.append(["ITEM001", "Sample Item 1", "All Item Groups", "Nos", "10", "100", "150", "Yes"])
        
        file_path = get_site_path("public", "files", "Bulk_Item_Import_Template.xlsx")
        wb.save(file_path)
        
        return "/files/Bulk_Item_Import_Template.xlsx"

def clean_val(val):
    if val is None or str(val).lower() == "none":
        return ""
    return str(val).strip()

def run_validation(docname):
    doc = frappe.get_doc("Bulk Item Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.item_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        
        errors = []
        item_groups = frappe.get_all("Item Group", pluck="name")
        uoms = frappe.get_all("UOM", pluck="name")
        
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            item_code = clean_val(row[0])
            item_group = clean_val(row[2])
            uom = clean_val(row[3])
            
            if not item_code:
                errors.append(f"Row {i}: Item Code is missing.")
            if item_group and item_group not in item_groups:
                errors.append(f"Row {i}: Item Group '{item_group}' does not exist.")
            if uom and uom not in uoms:
                errors.append(f"Row {i}: UOM '{uom}' does not exist.")

        if errors:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Validation Errors:\n" + "\n".join(errors))
        else:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", "✅ All rows validated successfully. Ready for processing.")
        frappe.db.commit()
        
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()

def run_processing(docname):
    doc = frappe.get_doc("Bulk Item Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.item_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        
        # Handle Zip images
        image_map = {}
        if doc.image_zip:
            zip_doc = frappe.get_doc("File", {"file_url": doc.image_zip})
            zip_path = zip_doc.get_full_path()
            extract_path = get_site_path("public", "files", f"temp_images_{doc.name}")
            
            if not os.path.exists(extract_path):
                os.makedirs(extract_path)
            
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_path)
                for filename in os.listdir(extract_path):
                    code = os.path.splitext(filename)[0]
                    image_map[code] = os.path.join(extract_path, filename)

        summary = []
        created_count = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item_code = clean_val(row[0])
            if not item_code: continue
            
            try:
                if frappe.db.exists("Item", item_code):
                    item = frappe.get_doc("Item", item_code)
                else:
                    item = frappe.new_doc("Item")
                    item.item_code = item_code
                
                item.item_name = clean_val(row[1])
                item.item_group = clean_val(row[2]) or "All Item Groups"
                item.stock_uom = clean_val(row[3]) or "Nos"
                item.is_stock_item = 1 if clean_val(row[7]).lower() == "yes" else 0
                item.valuation_rate = flt(row[5])
                item.standard_rate = flt(row[6])
                
                # Link Image from Zip
                if item_code in image_map:
                    img_path = image_map[item_code]
                    with open(img_path, "rb") as f:
                        file_content = f.read()
                        saved_file = save_file(os.path.basename(img_path), file_content, "Item", item_code, is_private=0)
                        item.image = saved_file.file_url
                
                item.save(ignore_permissions=True)
                frappe.db.commit()
                
                summary.append(f"✅ {item_code}: Created/Updated successfully")
                created_count += 1
                
            except Exception as e:
                frappe.db.rollback()
                summary.append(f"❌ {item_code}: {str(e)}")

        doc.db_set("status", "Completed")
        doc.db_set("processing_log", f"SUMMARY (Total Success: {created_count}):\n" + "\n".join(summary))
        frappe.db.commit()

        # Cleanup temp images
        if doc.image_zip:
            import shutil
            shutil.rmtree(extract_path, ignore_errors=True)

    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Critical Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
