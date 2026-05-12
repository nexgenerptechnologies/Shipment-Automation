import frappe
from frappe.model.document import Document
from frappe.utils import flt, getdate, nowdate
import openpyxl
from io import BytesIO
import datetime


@frappe.whitelist()
def download_template():
    """Generates and downloads the Bulk PO Import Excel template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk PO Import Template"
    
    headers = [
        "PO Number", "PO Date", "Supplier Name", "Item Code", "Item Name", 
        "Description", "Quantity", "Rate", "Required By Date", "Line Number"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    frappe.response['filename'] = "Bulk_PO_Import_Template.xlsx"
    frappe.response['filecontent'] = output.getvalue()
    frappe.response['type'] = 'binary'


class BulkPOImport(Document):

    @frappe.whitelist()
    def start_validation(self):
        self.db_set("status", "Validating")
        self.db_set("validation_log", "⏳ Validation in progress. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_po_import.bulk_po_import.run_validation",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Validation started. Refresh in a few seconds."

    @frappe.whitelist()
    def start_creation(self):
        if self.status != "Validated":
            frappe.throw("Please validate the data first.")
        self.db_set("status", "Processing")
        self.db_set("processing_log", "⏳ Creating Purchase Orders. Please refresh in a few seconds.")
        frappe.db.commit()
        frappe.enqueue(
            "shipment_automation.shipment_automation.doctype.bulk_po_import.bulk_po_import.run_processing",
            queue="long", timeout=3600, docname=self.name,
        )
        return "Processing started. Refresh in a few seconds."


def get_column_map(sheet):
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    expected = {
        "po_num": ["PO Number", "Purchase Order Number"],
        "po_date": ["PO Date", "Purchase Order Date"],
        "supplier": ["Supplier Name", "Supplier"],
        "item_code": ["Item Code"],
        "item_name": ["Item Name"],
        "description": ["Description"],
        "quantity": ["Quantity", "Qty"],
        "rate": ["Rate"],
        "req_date": ["Required By Date", "Target Date"],
        "line_number": ["Line Number", "Line #"]
    }
    for idx, cell in enumerate(header_row):
        if not cell: continue
        clean = str(cell).strip().lower()
        for key, aliases in expected.items():
            if any(alias.lower() == clean for alias in aliases):
                mapping[key] = idx
    return mapping


def parse_excel_date(date_val):
    """STRICTLY forces DD/MM/YYYY parsing for Bulk PO Import."""
    if not date_val:
        return None
    
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        if date_val.day <= 12:
             try:
                 return datetime.date(date_val.year, date_val.day, date_val.month).strftime("%Y-%m-%d")
             except ValueError:
                 return date_val.strftime("%Y-%m-%d")
        return date_val.strftime("%Y-%m-%d")

    if isinstance(date_val, str):
        date_str = date_val.strip()
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"]:
            try:
                dt = datetime.datetime.strptime(date_str, fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
    try:
        res = getdate(date_val)
        return res.strftime("%Y-%m-%d") if res else None
    except:
        return None


def run_validation(docname):
    doc = frappe.get_doc("Bulk PO Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.po_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        errors = []
        ok_rows = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            
            po_num = str(row[col_map["po_num"]]).strip() if col_map.get("po_num") is not None else ""
            raw_po_date = row[col_map["po_date"]] if col_map.get("po_date") is not None else None
            supplier = str(row[col_map["supplier"]]).strip() if col_map.get("supplier") is not None else ""
            raw_req_date = row[col_map["req_date"]] if col_map.get("req_date") is not None else None
            line_val = str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""

            # ── Use Forced DD/MM/YYYY Parser ──
            po_date = parse_excel_date(raw_po_date)
            req_date = parse_excel_date(raw_req_date)

            if not po_num: errors.append(f"Row {row_idx}: PO Number missing.")
            if not po_date: errors.append(f"Row {row_idx}: PO Date missing.")
            if not supplier: errors.append(f"Row {row_idx}: Supplier missing.")
            
            if po_date and (getdate(po_date).year < 2000 or getdate(po_date).year > 2100):
                errors.append(f"Row {row_idx}: Invalid PO Date (Must be 2000-2100).")
            
            if po_date and req_date and getdate(po_date) > getdate(req_date):
                errors.append(f"Row {row_idx}: PO Date cannot be after Required By Date.")

            if line_val:
                import re
                match = re.search(r'(\d+)$', po_num)
                if match:
                    prefix = match.group(1)
                    if not line_val.startswith(f"{prefix}-"):
                         errors.append(f"Row {row_idx}: Line Number {line_val} must start with {prefix}-")

            if not any([f"Row {row_idx}:" in e for e in errors]):
                ok_rows += 1

        if not errors:
            doc.db_set("status", "Validated")
            doc.db_set("validation_log", f"✅ {ok_rows} row(s) validated successfully.")
        else:
            doc.db_set("status", "Failed")
            doc.db_set("validation_log", "❌ Issues found:\n\n" + "\n".join(errors))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("validation_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()


def run_processing(docname):
    doc = frappe.get_doc("Bulk PO Import", docname)
    try:
        file_doc = frappe.get_doc("File", {"file_url": doc.po_excel})
        wb = openpyxl.load_workbook(file_doc.get_full_path(), data_only=True)
        sheet = wb.active
        col_map = get_column_map(sheet)
        
        po_groups = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            po_id = str(row[col_map["po_num"]]).strip()
            if po_id not in po_groups:
                po_groups[po_id] = []
            po_groups[po_id].append(row)

        created_pos = []
        for po_id, rows in po_groups.items():
            if frappe.db.exists("Purchase Order", po_id):
                continue

            first_row = rows[0]
            supplier = str(first_row[col_map["supplier"]]).strip()
            po_date = parse_excel_date(first_row[col_map["po_date"]])
            
            s_details = frappe.db.get_value("Supplier", supplier, ["default_currency"], as_dict=True)
            
            po = frappe.new_doc("Purchase Order")
            po.name = po_id
            
            available_series = frappe.get_meta("Purchase Order").get_field("naming_series").options.split("\n")
            for series in available_series:
                prefix = series.replace(".####", "").replace(".YY.", "").replace(".YYYY.", "").strip()
                if prefix and po_id.startswith(prefix):
                    po.naming_series = series
                    break

            po.supplier = supplier
            po.transaction_date = po_date
            po.company = frappe.db.get_single_value("Global Defaults", "default_company")
            po.currency = (s_details.default_currency if s_details else None) or frappe.db.get_single_value("Global Defaults", "default_currency")
            
            for row in rows:
                qty = flt(row[col_map["quantity"]])
                rate = flt(row[col_map["rate"]])
                req_date = parse_excel_date(row[col_map["req_date"]])
                line_val = str(row[col_map["line_number"]]).strip() if col_map.get("line_number") is not None and row[col_map["line_number"]] else ""
                
                item = po.append("items", {
                    "item_code": str(row[col_map["item_code"]]).strip(),
                    "qty": qty,
                    "rate": rate,
                    "schedule_date": req_date or po_date
                })
                item.run_method("set_missing_values")
                
                l_field = "line_number"
                if not hasattr(item, "line_number") and hasattr(item, "custom_line_number"):
                    l_field = "custom_line_number"
                if line_val: setattr(item, l_field, line_val)

            po.run_method("set_missing_values")
            po.run_method("calculate_taxes_and_totals")
            
            po.flags.ignore_permissions = True
            po.db_insert()
            for child in po.get_all_children():
                child.db_insert()
            po.run_method("on_update")
            
            created_pos.append(po.name)

        doc.db_set("status", "Completed")
        doc.db_set("processing_log", "SUMMARY:\n" + "\n".join([f"✅ {name}" for name in created_pos]))
        frappe.db.commit()
    except Exception:
        doc.db_set("status", "Failed")
        doc.db_set("processing_log", f"❌ Error:\n{frappe.get_traceback()}")
        frappe.db.commit()
